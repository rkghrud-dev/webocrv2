from __future__ import annotations

import argparse
import base64
from contextlib import contextmanager, nullcontext
import csv
import hashlib
import hmac
import json
import mimetypes
import os
import queue
import re
import ssl
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
import urllib.parse
import urllib.request
import urllib.error


ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT.parent
BACKEND_ROOT = PROJECT_ROOT / "backend"
BRIDGE_SCRIPT = PROJECT_ROOT / "KeywordOcr.App" / "Bridge" / "run_pipeline_bridge.py"
DATA_ROOT = ROOT / "data"
UPLOAD_ROOT = DATA_ROOT / "uploads"
LOGO_ROOT = DATA_ROOT / "logos"
JOBS_ROOT = DATA_ROOT / "jobs"
SEED_ROOT = DATA_ROOT / "seeds"
EXPORT_ROOT = DATA_ROOT / "exports"
MARKET_KEY_ROOT = DATA_ROOT / "market_keys"
MARKET_KEY_SETTINGS = MARKET_KEY_ROOT / "settings.json"
DOTNET_UPLOAD_PROJECT = PROJECT_ROOT / "KeywordOcr.App.Tests" / "KeywordOcr.App.Tests.csproj"
DESKTOP_KEY_ROOT = Path.home() / "Desktop" / "key"

for directory in (UPLOAD_ROOT, LOGO_ROOT, JOBS_ROOT, SEED_ROOT, EXPORT_ROOT, MARKET_KEY_ROOT):
    directory.mkdir(parents=True, exist_ok=True)

ACTIVE_PROCESSES: dict[str, subprocess.Popen] = {}
ACTIVE_PROCESS_LOCK = threading.Lock()
CANCELLED_JOB_IDS: set[str] = set()
MARKET_KEY_OVERLAY_LOCK = threading.Lock()
ACTIVE_MARKET_UPLOAD_LOCK = threading.Lock()
MARKET_UPLOAD_TIMEOUT_SECONDS = 900


KEYWORD_POOL_CATEGORIES = [
    {
        "id": "identity",
        "label": "상품 정체성",
        "description": "상품명, 표준명, 카테고리성 단어",
    },
    {
        "id": "function",
        "label": "기능",
        "description": "고정, 보강, 방지, 연결, 수선 등",
    },
    {
        "id": "usePlace",
        "label": "사용처",
        "description": "신발, 콘센트함, 스위치 박스, 가구 등",
    },
    {
        "id": "problemSolving",
        "label": "문제 해결",
        "description": "미끄럼방지, 누수방지, 흔들림방지 등",
    },
    {
        "id": "materialSpec",
        "label": "재질/식별 규격",
        "description": "EVA, ABS, PA66, M8, 86형 등",
    },
    {
        "id": "userSituation",
        "label": "사용자/상황",
        "description": "DIY, 수리, 시공 등",
    },
    {
        "id": "synonyms",
        "label": "동의어/다른 명칭",
        "description": "현장명, 별칭, 다른 표기",
    },
]

SALES_MARKETS = ["네이버", "쿠팡", "롯데ON", "11번가", "ESM"]
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}

COMPOUND_SPACING_RULES = [
    ("카라비너릴고리", "카라비너 릴고리"),
    ("와이어릴고리", "와이어 릴고리"),
    ("릴고리카라비너", "릴고리 카라비너"),
    ("카라비너릴홀더", "카라비너 릴홀더"),
    ("릴홀더와이어", "릴홀더 와이어"),
    ("와이어릴홀더", "와이어 릴홀더"),
    ("카라비너클립", "카라비너 클립"),
    ("백팩카라비너", "백팩 카라비너"),
    ("ABS카라비너", "ABS 카라비너"),
    ("카라비너고리", "카라비너 고리"),
    ("카라비너열쇠고리", "카라비너 열쇠고리"),
    ("카라비너가방고리", "카라비너 가방고리"),
    ("카라비너와이어", "카라비너 와이어"),
    ("키고리릴", "키고리 릴"),
    ("쿠션깔창", "쿠션 깔창"),
    ("신발깔창", "신발 깔창"),
    ("소프트깔창", "소프트 깔창"),
    ("쿠션인솔", "쿠션 인솔"),
    ("보강패드", "보강 패드"),
    ("충격완화", "충격 완화"),
    ("미끄럼방지", "미끄럼 방지"),
    ("먼지방지", "먼지 방지"),
    ("방수방진", "방수 방진"),
    ("열기차단", "열기 차단"),
    ("햇빛차단", "햇빛 차단"),
    ("사계절보호", "사계절 보호"),
]

BANNED_MARKETING_TERMS = [
    "발편한", "발 편한", "편한발", "편한 발",
    "무료배송", "할인", "세일", "특가", "추천", "인기", "베스트", "핫템",
    "가성비", "저렴한", "예쁜", "프리미엄", "고급", "고급형", "최고급",
    "판매", "가격", "문의", "상담",
]

LISTING_IMAGE_COLUMNS = [
    "이미지등록(목록)",
    "이미지등록(상세)",
    "목록이미지",
    "대표이미지",
    "대표 이미지",
    "상품이미지",
    "상품 이미지",
    "이미지",
    "썸네일",
]

ADDITIONAL_IMAGE_COLUMNS = [
    "이미지등록(추가)",
    "추가이미지",
    "추가 이미지",
    "권장이미지",
    "권장 이미지",
    "부가이미지",
]

DETAIL_IMAGE_COLUMNS = [
    "상품 상세설명",
    "모바일 상품 상세설명",
    "상세설명",
    "상품상세설명",
    "상세이미지",
]

DETAIL_HTML_COLUMNS = [
    "상품 상세설명",
    "모바일 상품 상세설명",
    "상세설명",
    "상품상세설명",
]

SEED_ANALYSIS_POLICY = {
    "purpose": "1차 시드는 원본 상품을 마켓별 작업 전에 정리하는 기준 데이터셋이다.",
    "keywordCategories": KEYWORD_POOL_CATEGORIES,
    "candidateOrder": [category["id"] for category in KEYWORD_POOL_CATEGORIES],
    "inputPriority": [
        "원본 상품명과 GS코드",
        "원본 옵션 컬럼과 optionItems",
        "OCR/사진 분석에서 상품과 직접 연결되는 사실",
        "이전 V파이프라인 키워드 후보",
    ],
    "searchTermRule": {
        "productName": "상품명은 표준명과 대표 검색어 중심으로 정확하게 만든다.",
        "searchTerms": "검색어설정은 표준어, 현장명, 별칭을 넓게 섞되 무관 인기어와 일부러 만든 오타는 넣지 않는다.",
        "marketSplit": "Cafe24는 실제 판매 채널이 아니라 네이버/쿠팡/ESM/11번가/롯데ON으로 나누기 전의 공통 키워드 풀로 본다.",
        "synonymExamples": {
            "나사": ["피스", "볼트", "체결나사"],
            "앵커": ["앙카", "칼블럭", "벽고정"],
            "육각렌치": ["알렌키", "L렌치", "렌치"],
            "소켓": ["복스알", "복스", "소켓렌치"],
            "라쳇핸들": ["깔깔이", "라쳇렌치"],
            "몽키스패너": ["몽키", "조절렌치"],
            "타정기": ["타카", "에어타카"],
            "그라인더": ["핸드그라인더", "절단기"],
            "드라이버": ["도라이바", "십자드라이버", "일자드라이버"],
            "콘센트함 고정나사": ["항공나사", "고정핀", "스위치박스 나사"],
        },
    },
    "sizeQuantityRule": {
        "quantity": "수량/구성은 검색어에 유지 가능하다. 예: 100p, 300p, 10개, 3세트, 5매, 2입",
        "singleSpec": "단일 규격 사이즈와 상품 식별 규격은 유지 가능하다. 예: 1M, 2mm, 35mm, M8, 86형, PA66, 304, ABS, EVA",
        "optionSize": "옵션형 상품의 숫자/색상은 원본 옵션 컬럼을 기준으로 판단한다. OCR에서만 나온 숫자 규격은 상품명/검색어/태그에서 제외한다.",
        "optionRange": "옵션 숫자가 230/240/250/260/270처럼 명확하면 필요 시 230-270 사이즈 선택형처럼 압축하고, 전체 옵션값을 나열하지 않는다.",
        "colorOption": "색상 옵션은 상품명에 넣지 않고 옵션 컬럼에서 처리한다.",
    },
    "rejectRule": {
        "ocrNoise": "OCR 눈금, 이미지 배경 숫자, 배송/상담/주소/전화/판매자 안내 문구는 후보에서 제외한다.",
        "salesNoise": "무료배송, 급배송, 할인, 추천, 가격, 인기, 베스트, 문의, 상담 등 판매 문구는 제외한다.",
        "duplication": "같은 뜻의 단어를 반복하거나 같은 구를 두 번 붙인 후보는 제외한다.",
    },
    "imageRule": {
        "listingSize": "대표/추가 이미지는 1000x1000 기준으로 가공한다.",
        "representativeDefault": "별도 선택 전에는 1번 이미지를 대표이미지로 둔다.",
    },
}


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_name(name: str, fallback: str = "upload.bin") -> str:
    base = Path(name or fallback).name
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", base).strip(" .")
    return base or fallback


def read_json(path: Path, default: dict | None = None) -> dict:
    if not path.exists():
        return default or {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default or {}


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def is_within(parent: Path, child: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def runtime_web_url(path: Path) -> str:
    resolved = path.resolve()
    if not is_within(ROOT, resolved):
        return str(resolved)
    relative = resolved.relative_to(ROOT).as_posix()
    return "/" + urllib.parse.quote(relative, safe="/:._-")


def image_file_sort_key(path: Path) -> tuple[int, str]:
    match = re.search(r"_(\d+)(?:\.[^.]+)$", path.name)
    if match:
        return int(match.group(1)), path.name.lower()
    numbers = re.findall(r"\d+", path.stem)
    return int(numbers[-1]) if numbers else 9999, path.name.lower()


def resolve_seed_path(value: str) -> Path:
    raw = text_value(value)
    if not raw:
        raise ValueError("seed path missing")
    path = Path(raw)
    if not path.is_absolute():
        path = SEED_ROOT / safe_name(raw)
    path = path.resolve()
    if not is_within(SEED_ROOT, path):
        raise ValueError("seed path outside data/seeds")
    if path.suffix.lower() != ".json" or ".webseed" not in path.name.lower():
        raise ValueError("not a webseed json file")
    return path


def seed_summary(path: Path) -> dict:
    payload = hydrate_seed_payload(read_json(path, {}))
    products = payload.get("products") if isinstance(payload.get("products"), list) else []
    first = products[0] if products else {}
    images = first.get("images", {}) if isinstance(first, dict) else {}
    return {
        "id": path.stem,
        "name": path.name,
        "createdAt": payload.get("createdAt") or datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "rows": payload.get("sourceFilter", {}).get("filteredRows", len(products)) if isinstance(payload.get("sourceFilter"), dict) else len(products),
        "gsCodes": len(products),
        "size": path.stat().st_size,
        "thumbnail": images.get("representative") or images.get("sourceThumb") or "",
        "path": str(path),
    }


def list_seed_summaries() -> list[dict]:
    items = []
    for path in sorted(SEED_ROOT.glob("*.webseed.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            items.append(seed_summary(path))
        except Exception:
            continue
    return items


def market_key_id(account: str, market: str) -> str:
    return f"{text_value(account).upper()}:{text_value(market)}"


def read_market_key_settings() -> dict:
    payload = read_json(MARKET_KEY_SETTINGS, {"items": {}})
    if not isinstance(payload.get("items"), dict):
        payload["items"] = {}
    return payload


def write_market_key_settings(payload: dict) -> None:
    write_json(MARKET_KEY_SETTINGS, payload)


def market_key_summaries() -> list[dict]:
    payload = read_market_key_settings()
    items: list[dict] = []
    for key, item in payload.get("items", {}).items():
        path = Path(item.get("path", ""))
        exists = path.is_file() and is_within(MARKET_KEY_ROOT, path)
        items.append({
            "key": key,
            "account": item.get("account", ""),
            "market": item.get("market", ""),
            "mode": item.get("mode", "key"),
            "fileName": item.get("fileName", ""),
            "size": path.stat().st_size if exists else 0,
            "updatedAt": item.get("updatedAt", ""),
            "exists": exists,
        })
    return sorted(items, key=lambda item: (item.get("market", ""), item.get("account", "")))


def read_secret_payload(path: Path) -> dict[str, str]:
    text = path.read_text(encoding="utf-8-sig", errors="replace").strip()
    out: dict[str, str] = {}
    try:
        raw = json.loads(text)
        if isinstance(raw, dict):
            for key, value in raw.items():
                if isinstance(value, (str, int, float, bool)):
                    out[str(key)] = str(value)
                elif isinstance(value, dict):
                    for child_key, child_value in value.items():
                        if isinstance(child_value, (str, int, float, bool)):
                            out[str(child_key)] = str(child_value)
                            out[f"{key}.{child_key}"] = str(child_value)
    except Exception:
        pass
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        out[key.strip()] = value.strip().strip('"').strip("'")
    if not out and text and len(text) < 400:
        out["value"] = text
    return out


def pick_secret(payload: dict[str, str], *names: str) -> str:
    lower = {str(key).lower(): value for key, value in payload.items()}
    for name in names:
        if payload.get(name):
            return payload[name]
        value = lower.get(name.lower())
        if value:
            return value
    return ""


def scrub_secret(value: object, limit: int = 180) -> str:
    text = str(value or "")
    text = re.sub(r"(?i)(access[_-]?token|refresh[_-]?token|client[_-]?secret|secret[_-]?key|api[_-]?key|signature)[^,}\n]*", r"\1=***", text)
    text = re.sub(r"Bearer\s+[A-Za-z0-9._~+/=-]+", "Bearer ***", text)
    return text[:limit]


def request_text(method: str, url: str, headers: dict | None = None, body: bytes | None = None, timeout: int = 20) -> tuple[bool, int | str, str]:
    req = urllib.request.Request(url, data=body, headers=headers or {})
    req.get_method = lambda: method
    ctx = ssl.create_default_context()
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=timeout) as resp:
            return True, resp.status, resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        return False, exc.code, exc.read().decode("utf-8", errors="replace")
    except Exception as exc:
        return False, "ERR", str(exc)


def set_secret_alias(raw: dict, aliases: list[str], value: str) -> None:
    for alias in aliases:
        if alias in raw:
            raw[alias] = value
            return
    raw[aliases[0]] = value


def refresh_cafe24_access_token(path: Path, cfg: dict[str, str], mall_id: str) -> tuple[bool, str]:
    refresh_token = pick_secret(cfg, "REFRESH_TOKEN", "refresh_token", "refreshToken")
    client_id = pick_secret(cfg, "CLIENT_ID", "client_id", "clientId")
    client_secret = pick_secret(cfg, "CLIENT_SECRET", "client_secret", "clientSecret")
    if not refresh_token or not client_id or not client_secret:
        return False, "refresh_token/client 정보 없음"
    basic = base64.b64encode(f"{client_id}:{client_secret}".encode("utf-8")).decode("ascii")
    body = urllib.parse.urlencode({
        "grant_type": "refresh_token",
        "refresh_token": refresh_token,
    }).encode("utf-8")
    ok, status, response = request_text("POST", f"https://{mall_id}.cafe24api.com/api/v2/oauth/token", {
        "Authorization": f"Basic {basic}",
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
    }, body)
    if not ok:
        return False, f"토큰 갱신 실패 status={status}: {scrub_secret(response)}"
    try:
        token_payload = json.loads(response)
    except Exception:
        return False, "토큰 갱신 응답 파싱 실패"
    access_token = token_payload.get("access_token")
    if not access_token:
        return False, "토큰 갱신 응답에 access_token 없음"
    try:
        raw = json.loads(path.read_text(encoding="utf-8-sig", errors="replace"))
        if isinstance(raw, dict):
            set_secret_alias(raw, ["AccessToken", "ACCESS_TOKEN", "access_token"], access_token)
            if token_payload.get("refresh_token"):
                set_secret_alias(raw, ["RefreshToken", "REFRESH_TOKEN", "refresh_token"], token_payload["refresh_token"])
            set_secret_alias(raw, ["UpdatedAt", "UPDATED_AT", "updated_at"], now_text())
            path.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        return True, access_token
    return True, access_token


def test_cafe24_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    mall_id = pick_secret(cfg, "MALL_ID", "mall_id", "mallId")
    token = pick_secret(cfg, "ACCESS_TOKEN", "access_token", "accessToken")
    api_version = pick_secret(cfg, "API_VERSION", "api_version", "apiVersion") or "2025-12-01"
    shop_no = pick_secret(cfg, "SHOP_NO", "shop_no", "shopNo") or "1"
    if not mall_id or not token:
        return False, "CONFIG", "MALL_ID/ACCESS_TOKEN 없음"
    query = urllib.parse.urlencode({"shop_no": shop_no, "limit": 1})
    url = f"https://{mall_id}.cafe24api.com/api/v2/admin/products?{query}"
    ok, status, body = request_text("GET", url, {
        "Authorization": f"Bearer {token}",
        "X-Cafe24-Api-Version": api_version,
        "Accept": "application/json",
    })
    if not ok and status == 401:
        refreshed, new_token_or_message = refresh_cafe24_access_token(path, cfg, mall_id)
        if refreshed:
            ok, status, body = request_text("GET", url, {
                "Authorization": f"Bearer {new_token_or_message}",
                "X-Cafe24-Api-Version": api_version,
                "Accept": "application/json",
            })
            return ok, status, "token 갱신 후 products 조회 OK" if ok else scrub_secret(body)
        return False, status, new_token_or_message
    return ok, status, "products 조회 OK" if ok else scrub_secret(body)


def test_naver_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    client_id = pick_secret(cfg, "NAVER_COMMERCE_CLIENT_ID", "CLIENT_ID", "client_id")
    client_secret = pick_secret(cfg, "NAVER_COMMERCE_CLIENT_SECRET", "CLIENT_SECRET", "client_secret")
    if not client_id or not client_secret:
        return False, "CONFIG", "CLIENT_ID/CLIENT_SECRET 없음"
    try:
        import bcrypt
    except Exception as exc:
        return False, "LOCAL", f"bcrypt 모듈 없음: {exc}"
    timestamp = int((time.time() - 3) * 1000)
    sign = base64.standard_b64encode(
        bcrypt.hashpw(f"{client_id}_{timestamp}".encode("utf-8"), client_secret.encode("utf-8"))
    ).decode("utf-8")
    body = urllib.parse.urlencode({
        "client_id": client_id,
        "timestamp": timestamp,
        "client_secret_sign": sign,
        "grant_type": "client_credentials",
        "type": "SELF",
    }).encode("utf-8")
    ok, status, response = request_text("POST", "https://api.commerce.naver.com/external/v1/oauth2/token", {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
    }, body)
    return ok, status, "token 발급 OK" if ok else scrub_secret(response)


def coupang_auth(access_key: str, secret_key: str, method: str, path: str, query: str = "") -> str:
    stamp = datetime.utcnow().strftime("%y%m%dT%H%M%SZ")
    message = stamp + method + path + query
    signature = hmac.new(secret_key.encode("utf-8"), message.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"CEA algorithm=HmacSHA256, access-key={access_key}, signed-date={stamp}, signature={signature}"


def test_coupang_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    access_key = pick_secret(cfg, "access_key", "ACCESS_KEY", "COUPANG_ACCESS_KEY")
    secret_key = pick_secret(cfg, "secret_key", "SECRET_KEY", "COUPANG_SECRET_KEY")
    if not access_key or not secret_key:
        return False, "CONFIG", "access_key/secret_key 없음"
    api_path = "/v2/providers/seller_api/apis/api/v1/marketplace/meta/category-related-metas/display-category-codes/56137"
    ok, status, body = request_text("GET", "https://api-gateway.coupang.com" + api_path, {
        "Authorization": coupang_auth(access_key, secret_key, "GET", api_path),
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json",
    })
    return ok, status, "category meta 조회 OK" if ok else scrub_secret(body)


def test_lotteon_key(path: Path) -> tuple[bool, int | str, str]:
    cfg = read_secret_payload(path)
    api_key = pick_secret(cfg, "api_key", "LOTTEON_API_KEY", "ApiKey", "apiKey", "value")
    if not api_key:
        return False, "CONFIG", "LOTTEON_API_KEY/api_key 없음"
    ok, status, body = request_text("GET", "https://openapi.lotteon.com/v1/openapi/common/v1/identity", {
        "Authorization": f"Bearer {api_key}",
        "Accept": "application/json",
    })
    return ok, status, "identity 조회 OK" if ok else scrub_secret(body)


def test_market_key(item: dict) -> dict:
    path = Path(item.get("path", ""))
    market = item.get("market", "")
    result = {
        "key": market_key_id(item.get("account", ""), market),
        "account": item.get("account", ""),
        "market": market,
        "fileName": item.get("fileName", ""),
        "ok": False,
        "status": "SKIP",
        "message": "",
        "testedAt": now_text(),
    }
    if not path.is_file() or not is_within(MARKET_KEY_ROOT, path):
        result.update({"status": "MISSING", "message": "저장된 파일 없음"})
        return result
    testers = {
        "Cafe24": test_cafe24_key,
        "네이버": test_naver_key,
        "쿠팡": test_coupang_key,
        "롯데ON": test_lotteon_key,
    }
    tester = testers.get(market)
    if not tester:
        result.update({"status": "SKIP", "message": "엑셀 서식 마켓"})
        return result
    ok, status, message = tester(path)
    result.update({"ok": ok, "status": status, "message": message})
    return result


def infer_seed_progress(line: str, fallback: int = 12) -> tuple[int, str, str]:
    clean = text_value(line)
    progress = fallback
    stage = "파이프라인 실행 중"
    current_gs = ""
    gs_match = re.search(r"(GS\d{7}[A-Z]?)", clean, re.IGNORECASE)
    if gs_match:
        current_gs = gs_match.group(1).upper()
    count_match = re.search(r"\[(\d+)\s*/\s*(\d+)\]", clean)
    if count_match:
        done = int(count_match.group(1))
        total = max(1, int(count_match.group(2)))
        progress = max(progress, min(86, 18 + int((done / total) * 60)))
    lowered = clean.lower()
    if "다운로드" in clean or "download" in lowered:
        stage = "이미지 다운로드"
        progress = max(progress, 24)
    if "ocr" in lowered:
        stage = "OCR 분석"
        progress = max(progress, 45)
    if "gpt" in lowered or "사진" in clean or "분석" in clean:
        stage = "사진/상품 분석"
        progress = max(progress, 58)
    if "엑셀" in clean or "excel" in lowered:
        stage = "기본 데이터 정리"
        progress = max(progress, 72)
    if clean.startswith("__RESULT__"):
        stage = "결과 파일 수집"
        progress = 88
    return progress, stage, current_gs


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def text_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def price_value(value: object) -> int:
    text = text_value(value)
    text = re.sub(r"[^\d.-]", "", text)
    if not text:
        return 0
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


def v3_price_multiplier(value: object) -> float:
    try:
        amount = float(value)
    except Exception:
        return 2.0
    if amount >= 20000:
        return 1.6
    if amount >= 10000:
        return 1.8
    return 2.0


def round_100(value: object) -> int:
    try:
        return int(round(float(value), -2))
    except Exception:
        return 0


def source_price_info(row: dict, headers: dict[str, str]) -> dict[str, int]:
    supply = first_price(row, headers, ["공급가", "공급 가격", "공급가격", "매입가"])
    raw_sale = first_price(row, headers, ["판매가", "상품가", "가격"])
    raw_consumer = first_price(row, headers, ["소비자가", "정가"])
    if supply:
        sale = round_100(supply * v3_price_multiplier(supply))
        consumer = round_100(sale * 1.2)
    else:
        sale = raw_sale or raw_consumer
        consumer = raw_consumer or (round_100(sale * 1.2) if sale else 0)
    return {
        "supplyPrice": supply,
        "salePrice": sale,
        "consumerPrice": consumer,
    }


def first_value(row: dict, headers: dict[str, str], names: list[str]) -> str:
    for name in names:
        key = headers.get(normalize_header(name))
        if key is not None:
            value = text_value(row.get(key, ""))
            if value:
                return value
    return ""


def first_price(row: dict, headers: dict[str, str], names: list[str]) -> int:
    for name in names:
        key = headers.get(normalize_header(name))
        if key is not None:
            value = price_value(row.get(key, ""))
            if value:
                return value
    return 0


GS_CODE_RE = re.compile(r"(GS\d{7})([A-Z0-9]*)", re.IGNORECASE)


def extract_gs(*values: object) -> str:
    for value in values:
        match = GS_CODE_RE.search(text_value(value))
        if match:
            return match.group(0).upper()
    return ""


def split_gs(gs: str) -> tuple[str, str]:
    match = GS_CODE_RE.search(gs or "")
    if not match:
        return gs.upper(), ""
    return match.group(1).upper(), (match.group(2) or "").upper()


def short_name(name: str, gs: str) -> str:
    text = GS_CODE_RE.sub("", name or "")
    text = re.sub(r"\s+", " ", text).strip(" -_/|")
    return text or name or gs


def common_word_prefix(values: list[str]) -> str:
    token_sets = [value.split() for value in values if value]
    if len(token_sets) < 2:
        return ""
    prefix: list[str] = []
    for tokens in zip(*token_sets):
        if len(set(tokens)) != 1:
            break
        prefix.append(tokens[0])
    return " ".join(prefix).strip()


def strip_option_label(name: str, base_name: str) -> str:
    if base_name and name.startswith(base_name):
        return name[len(base_name):].strip(" -_/|")
    return ""


def compact_option_summary(labels: list[str], total: int) -> str:
    labels = [label for label in labels if label and label != "단일"]
    if total <= 1:
        return labels[0] if labels else "단일"
    if not labels:
        return f"{total} 옵션"
    shown = " / ".join(labels[:8])
    more = f" 외 {len(labels) - 8}" if len(labels) > 8 else ""
    return f"{total} 옵션 · {shown}{more}"


def option_meta(option_summary: str, option_items: list[dict]) -> dict:
    summary = text_value(option_summary) or "단일"
    match = re.search(r"(\d+)\s*옵션", summary)
    summary_count = int(match.group(1)) if match else 0
    count = max(len(option_items or []), summary_count)
    has_options = count > 1 or (summary != "단일" and bool(re.search(r"옵션|/|,", summary)))
    return {
        "optionType": "option" if has_options else "single",
        "optionCount": max(count, 1) if has_options else 1,
    }


def empty_keyword_pool() -> dict[str, list[str]]:
    return {category["id"]: [] for category in KEYWORD_POOL_CATEGORIES}


TERM_STOPWORDS = {
    "상품", "제품", "옵션", "선택", "수량", "소재", "사이즈", "참고사항", "제조국", "중국",
    "수입사", "굿셀러스", "Product", "Profile", "SIZE", "Advantage", "000", "http", "https",
    "홈런마켓", "급배송", "무료", "묶음배송", "대량구매", "문의", "상담", "문자상담",
    "택배사", "방문수령", "경기도", "광주시", "친절하게", "환영합니다", "주의", "사항",
    "앞면", "뒷면", "참고용", "측정방법", "오차", "실제", "선택하세요",
}

CATEGORY_HINTS = {
    "identity": [
        "카라비너", "릴홀더", "릴고리", "고리", "라벨", "태그", "이름표", "커버", "브러시",
        "장갑", "스티커", "파우치", "클립", "후크", "나사", "브라켓", "정리함",
    ],
    "function": [
        "고정", "보강", "방지", "연결", "수선", "탈부착", "체결", "걸이", "걸어",
        "늘어나는", "간편", "빠르고", "장식", "보관", "수납", "보호", "충전", "거치",
    ],
    "usePlace": [
        "가방", "백팩", "텐트", "스트랩", "벨트", "하네스", "손전등", "열쇠", "신발",
        "콘센트", "스위치", "가구", "차량", "현장", "아웃도어", "일상", "캠핑",
    ],
    "problemSolving": [
        "미끄럼방지", "누수방지", "흔들림방지", "분실방지", "낙하방지", "방지",
        "스크래치", "보호", "오차", "편의", "간편",
    ],
    "userSituation": [
        "DIY", "수리", "시공", "작업", "현장", "아웃도어", "캠핑", "등산", "일상",
        "휴대", "여행", "장갑", "활동",
    ],
}

SYNONYM_SEEDS = {
    "카라비너": ["카라비너고리", "카라비너클립", "카라비너열쇠고리", "카라비너가방고리", "카라비너와이어", "백팩카라비너", "아웃도어카라비너"],
    "릴홀더": ["릴고리", "릴홀더와이어", "와이어릴홀더", "키릴", "키고리릴"],
    "릴고리": ["릴홀더", "와이어고리", "키고리릴"],
    "클립": ["집게", "고정클립", "후크"],
    "나사": ["피스", "볼트", "체결나사"],
    "앵커": ["앙카", "칼블럭", "벽고정"],
    "장갑": ["작업장갑", "현장장갑", "보호장갑"],
}


def unique_terms(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = re.sub(r"\s+", " ", text_value(value)).strip(" ,/|")
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def normalize_term(term: object) -> str:
    text = text_value(term)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"[{}\"'`]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,./|·:;_-")
    if not text or len(text) < 2:
        return ""
    if text in TERM_STOPWORDS or text.upper() in TERM_STOPWORDS:
        return ""
    if GS_CODE_RE.fullmatch(text):
        return ""
    if re.fullmatch(r"\d+", text):
        return ""
    if len(text) > 36:
        return ""
    return text


def apply_compound_spacing(value: object) -> str:
    text = text_value(value)
    for source, target in COMPOUND_SPACING_RULES:
        text = text.replace(source, target)
    return re.sub(r"\s+", " ", text).strip()


def remove_marketing_terms(value: object) -> str:
    text = apply_compound_spacing(value)
    for word in BANNED_MARKETING_TERMS:
        text = re.sub(re.escape(word), " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip(" ,/|")


def clean_product_title(value: object) -> str:
    text = remove_marketing_terms(value)
    text = re.sub(r"\b1\s*(?:개|입|매|p|P|pcs|PCS|pc|PC)\b", " ", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip(" ,/|")


def compact_search_keyword(value: object) -> str:
    text = remove_marketing_terms(value)
    text = re.sub(r"\s+", "", text)
    return text.strip(" ,/|")


def split_candidate_terms(*values: object) -> list[str]:
    terms: list[str] = []
    for value in values:
        raw = text_value(value)
        if not raw:
            continue
        raw = re.sub(r"<[^>]+>", " ", raw)
        for chunk in re.split(r"[,/|;·\n\r\t]+", raw):
            chunk = normalize_term(chunk)
            if not chunk:
                continue
            terms.append(chunk)
            words = [normalize_term(word) for word in re.split(r"\s+", chunk)]
            words = [word for word in words if word]
            terms.extend(words)
            for i in range(len(words) - 1):
                joined = normalize_term(f"{words[i]} {words[i + 1]}")
                if joined:
                    terms.append(joined)
    return unique_terms(terms)


NOTICE_COMMON_DEFAULTS = {
    "returnCostReason": "0",
    "noRefundReason": "0",
    "qualityAssuranceStandard": "0",
    "compensationProcedure": "0",
    "troubleShootingContents": "0",
}

NOTICE_REVIEW_DEFAULTS = {
    "warrantyPolicy": "0",
    "afterServiceDirector": "0",
    "caution": "0",
}

NOTICE_FIELD_LABELS = [
    "소재", "재질", "수량", "색상", "사이즈", "수입사", "수입원", "제조사",
    "제조자", "제조원", "제조국", "원산지", "A/S", "AS", "고객센터",
]

NOTICE_NOISE_PATTERNS = [
    re.compile(r"^(?:high|bullet|advantage|product\s*profile|size)$", re.IGNORECASE),
    re.compile(r"^(?:in/mm|on|off|zero)$", re.IGNORECASE),
    re.compile(r"^[{}\"']+$"),
    re.compile(r"^\d{1,4}$"),
    re.compile(r"^\d{1,4}\s*mm$", re.IGNORECASE),
]


def clean_notice_value(value: object, max_length: int = 160) -> str:
    text = text_value(value)
    text = re.sub(r"[{}\"`]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,./|·:-")
    text = re.split(
        r"\s+(?:홈런마켓|급배송|평일|택배사|구매대행|국내|모든\s*제품|상품/대량구매|대량구매|퀵서비스|방문수령)\b",
        text,
        maxsplit=1,
    )[0].strip(" ,./|·:-")
    return text[:max_length].strip()


def notice_is_noise_line(value: object) -> bool:
    text = clean_notice_value(value, max_length=80)
    if not text:
        return True
    return any(pattern.search(text) for pattern in NOTICE_NOISE_PATTERNS)


def ocr_notice_lines(raw_text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in text_value(raw_text).splitlines():
        line = clean_notice_value(raw_line, max_length=220)
        if line:
            lines.append(line)
    return lines


def extract_labeled_notice_value(lines: list[str], labels: list[str], stop_labels: list[str] | None = None) -> str:
    stop_labels = stop_labels or NOTICE_FIELD_LABELS
    label_pattern = "|".join(re.escape(label) for label in labels)
    stop_pattern = "|".join(re.escape(label) for label in stop_labels if label not in labels)
    joined = " ".join(lines)
    match = re.search(
        rf"(?:^|\s)(?:{label_pattern})\s*[:：]?\s*(.+?)(?=\s+(?:{stop_pattern})\s*[:：]?|\s+(?:SIZE|Product Profile|홈런마켓|급배송)\b|$)",
        joined,
        re.IGNORECASE,
    )
    if match:
        value = clean_notice_value(match.group(1))
        if value and not notice_is_noise_line(value):
            return value

    exact_pattern = re.compile(rf"^(?:{label_pattern})\s*[:：]?$", re.IGNORECASE)
    inline_pattern = re.compile(rf"^(?:{label_pattern})\s*[:：]?\s*(.+)$", re.IGNORECASE)
    stop_line_pattern = re.compile(rf"^(?:{'|'.join(re.escape(label) for label in stop_labels)})\b", re.IGNORECASE)
    for index, line in enumerate(lines):
        inline = inline_pattern.search(line)
        if inline:
            value = clean_notice_value(inline.group(1))
            if value and not notice_is_noise_line(value):
                return value
        if exact_pattern.search(line):
            for next_line in lines[index + 1:index + 4]:
                if stop_line_pattern.search(next_line):
                    break
                if not notice_is_noise_line(next_line):
                    return clean_notice_value(next_line)
    return ""


def normalize_notice_size(raw_size: str, option_items: list[dict]) -> str:
    option_labels = [
        clean_notice_value(item.get("option"))
        for item in option_items or []
        if clean_notice_value(item.get("option")) and clean_notice_value(item.get("option")) != "단일"
    ]
    if option_labels:
        return " / ".join(unique_terms(option_labels[:20]))

    text = clean_notice_value(raw_size, max_length=240)
    text = re.split(r"\s+-?\s*위\s*\d*종|\s+수입사|\s+제조국|\s+SIZE\b", text, maxsplit=1, flags=re.IGNORECASE)[0]
    numbers = re.findall(r"(?<!\d)(?:[A-Z]\s*)?(\d{2,4})(?!\d)", text, flags=re.IGNORECASE)
    shoe_sizes = [number for number in numbers if 150 <= int(number) <= 350]
    if shoe_sizes:
        return " / ".join(unique_terms(shoe_sizes))
    return clean_notice_value(text)


def extract_notice_dimensions(raw_text: str) -> list[str]:
    values = re.findall(r"\b[A-Z]?\s*\d{2,4}\s*-\s*\d{1,4}\s*mm\s*[xX]\s*\d{1,4}\s*mm\b", raw_text, flags=re.IGNORECASE)
    return unique_terms([clean_notice_value(value).replace(" x ", "X").replace("x", "X") for value in values])[:12]


def infer_notice_type(row: dict, extracted: dict[str, str], raw_text: str) -> tuple[str, str]:
    joined = " ".join([
        text_value(row.get("name")),
        text_value(row.get("opt")),
        text_value(raw_text),
    ])
    if re.search(r"깔창|인솔|신발|운동화|구두|슬리퍼|부츠", joined):
        return "SHOES", "shoes"
    if re.search(r"가방|백팩|파우치|숄더백|토트백", joined):
        return "BAG", "bag"
    if re.search(r"의류|티셔츠|셔츠|바지|자켓|재킷|점퍼|원피스|스커트", joined):
        return "WEAR", "wear"
    return "ETC", "etc"


def build_naver_provided_notice(row: dict, ocr_record: dict) -> dict:
    fields = ocr_record.get("fields", {}) if isinstance(ocr_record, dict) else {}
    raw_text = text_value(ocr_record.get("rawText")) or text_value(fields.get("OCR텍스트"))
    lines = ocr_notice_lines(raw_text)
    source_name = clean_product_title(row.get("name")) or clean_notice_value(fields.get("상품명")) or text_value(row.get("gs"))
    option_items = row.get("optionItems", []) if isinstance(row.get("optionItems"), list) else []

    extracted = {
        "itemName": source_name,
        "modelName": text_value(row.get("gs")) or text_value(row.get("baseGs")),
        "material": extract_labeled_notice_value(lines, ["소재", "재질"]),
        "quantity": extract_labeled_notice_value(lines, ["수량"]),
        "color": extract_labeled_notice_value(lines, ["색상"]),
        "size": normalize_notice_size(extract_labeled_notice_value(lines, ["사이즈"]), option_items),
        "importer": extract_labeled_notice_value(lines, ["수입사", "수입원"]),
        "manufacturer": extract_labeled_notice_value(lines, ["제조사", "제조자", "제조원"]),
        "origin": extract_labeled_notice_value(lines, ["제조국", "원산지"]),
        "customerServicePhoneNumber": extract_labeled_notice_value(lines, ["A/S", "AS", "고객센터"]),
    }
    dimensions = extract_notice_dimensions(raw_text)
    if dimensions:
        extracted["sizeDetail"] = " / ".join(dimensions)
    extracted = {key: value for key, value in extracted.items() if value}

    notice_type, object_key = infer_notice_type(row, extracted, raw_text)
    manufacturer = extracted.get("manufacturer") or extracted.get("importer") or "상품상세 참조"
    material = extracted.get("material") or "상품상세 참조"
    color = extracted.get("color") or "상품상세 참조"
    size = extracted.get("size") or extracted.get("sizeDetail") or "상품상세 참조"

    if notice_type == "SHOES":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            **NOTICE_REVIEW_DEFAULTS,
            "material": material,
            "color": color,
            "size": size,
            "height": "해당사항 없음",
            "manufacturer": manufacturer,
        }
    elif notice_type == "WEAR":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            **NOTICE_REVIEW_DEFAULTS,
            "material": material,
            "color": color,
            "size": size,
            "manufacturer": manufacturer,
        }
    elif notice_type == "BAG":
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            **NOTICE_REVIEW_DEFAULTS,
            "type": source_name,
            "material": material,
            "color": color,
            "size": size,
            "manufacturer": manufacturer,
        }
    else:
        detail = {
            **NOTICE_COMMON_DEFAULTS,
            "itemName": source_name,
            "modelName": extracted.get("modelName") or text_value(row.get("baseGs")) or source_name,
            "manufacturer": manufacturer,
            "customerServicePhoneNumber": extracted.get("customerServicePhoneNumber") or "상품상세 참조",
        }

    needs_review: list[str] = []
    if extracted.get("importer") and not extracted.get("manufacturer"):
        needs_review.append("수입사를 제조자/수입자로 임시 사용했습니다.")
    if not extracted.get("customerServicePhoneNumber"):
        needs_review.append("A/S 전화번호는 OCR에서 명확히 확인되지 않았습니다.")
    if notice_type != "ETC":
        needs_review.append("카테고리 API 연동 전까지 OCR 문구로 고시 상품군을 추정했습니다.")
    if extracted.get("sizeDetail") and extracted.get("size"):
        needs_review.append("상세 치수는 보조값으로만 보관하고 옵션 사이즈를 우선했습니다.")

    matched_fields = {
        key: extracted[key]
        for key in ("material", "quantity", "color", "size", "sizeDetail", "importer", "manufacturer", "origin")
        if extracted.get(key)
    }
    status = "matched" if matched_fields else "empty"
    if needs_review:
        status = "partial"

    return {
        "status": status,
        "source": "ocr_label_match",
        "productInfoProvidedNoticeType": notice_type,
        "objectKey": object_key,
        "productInfoProvidedNotice": {
            "productInfoProvidedNoticeType": notice_type,
            object_key: detail,
        },
        "extractedFields": extracted,
        "matchedFields": matched_fields,
        "needsReview": needs_review,
    }


def find_material_spec_terms(*values: object, include_free_numeric: bool = True) -> list[str]:
    text = " ".join(text_value(value) for value in values)
    patterns = [
        r"\b(?:ABS|EVA|PA66|PVC|PP|PE|PU|SUS|STS)\b",
        r"(?:에어메쉬|폴리에스테르|나일론|스판|면|패브릭|부직포|실리콘|고무)",
        r"\b(?:304|316)\b",
        r"\bM\d+(?:\.\d+)?\b",
        r"\b\d+\s?형\b",
    ]
    if include_free_numeric:
        patterns.append(r"\b\d+(?:\.\d+)?\s?(?:mm|cm|m|M)\b")
    terms: list[str] = []
    for pattern in patterns:
        terms.extend(match.group(0).replace(" ", "") for match in re.finditer(pattern, text, re.IGNORECASE))
    return unique_terms(terms)


def terms_by_hints(terms: list[str], hints: list[str]) -> list[str]:
    return unique_terms([term for term in terms if any(hint.lower() in term.lower() for hint in hints)])


def synonym_terms(terms: list[str]) -> list[str]:
    joined = " ".join(terms)
    out: list[str] = []
    for trigger, synonyms in SYNONYM_SEEDS.items():
        if trigger in joined:
            out.extend(synonyms)
    return unique_terms(out)


def seed_keyword_pool_from_product(row: dict, ocr_text: str = "", keyword_record: dict | None = None) -> dict[str, list[str]]:
    pool = empty_keyword_pool()
    option_terms = [item.get("option", "") for item in row.get("optionItems", []) if item.get("option") != "단일"]
    keyword_record = keyword_record or {}
    generated_product_names = keyword_record.get("productNames", [])
    generated_search_terms = keyword_record.get("searchTerms", [])
    generated_longtails = keyword_record.get("longtails", [])
    generated_debug_terms = keyword_record.get("debugTerms", [])
    all_terms = split_candidate_terms(
        row.get("name", ""),
        row.get("opt", ""),
        " ".join(option_terms),
        ocr_text,
        " ".join(generated_product_names),
        " ".join(generated_search_terms),
        " ".join(generated_longtails),
        " ".join(generated_debug_terms),
    )
    pool["identity"] = unique_terms(
        [row.get("name", "")]
        + generated_product_names[:4]
        + terms_by_hints(all_terms, CATEGORY_HINTS["identity"])
    )[:18]
    pool["function"] = terms_by_hints(all_terms, CATEGORY_HINTS["function"])[:20]
    pool["usePlace"] = terms_by_hints(all_terms, CATEGORY_HINTS["usePlace"])[:20]
    pool["problemSolving"] = terms_by_hints(all_terms, CATEGORY_HINTS["problemSolving"])[:14]
    # 옵션형 상품은 OCR 치수 눈금/이미지 배경 숫자를 상품 규격으로 오인하기 쉽다.
    # 숫자 규격은 옵션/원본명에서만 보고, OCR에서는 재질명 중심으로만 가져온다.
    pool["materialSpec"] = unique_terms(
        find_material_spec_terms(row.get("name", ""), row.get("opt", ""), " ".join(option_terms), include_free_numeric=True)
        + find_material_spec_terms(ocr_text, " ".join(all_terms), include_free_numeric=not bool(option_terms))
    )[:18]
    pool["userSituation"] = terms_by_hints(all_terms, CATEGORY_HINTS["userSituation"])[:16]
    pool["synonyms"] = unique_terms(generated_longtails + synonym_terms(all_terms))[:24]
    return pool


def read_pipeline_ocr_summary(output_file: str, output_root: str = "") -> dict:
    path = Path(output_file or "")
    summary = {
        "available": False,
        "workbook": str(path) if output_file else "",
        "sheets": [],
        "ocrByGs": {},
    }
    if not output_file or not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return summary

    def read_ocr_workbook(workbook_path: Path) -> bool:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            summary["available"] = True
            summary["workbook"] = str(workbook_path)
            summary["sheets"] = list(workbook.sheetnames)
            if "OCR결과" not in workbook.sheetnames:
                return False

            sheet = workbook["OCR결과"]
            rows = sheet.iter_rows(values_only=True)
            headers = [text_value(value) for value in next(rows, [])]
            gs_col = next((i for i, header in enumerate(headers) if "GS" in header.upper()), 0)
            for row in rows:
                values = {headers[i] or f"col{i + 1}": text_value(row[i] if i < len(row) else "") for i in range(len(headers))}
                gs = extract_gs(row[gs_col] if gs_col < len(row) else "", values)
                if not gs:
                    continue
                base_gs, _ = split_gs(gs)
                raw_text_parts = [
                    value for key, value in values.items()
                    if value and not re.search(r"(이미지|image|path|경로|url)", key, re.IGNORECASE)
                ]
                summary["ocrByGs"][base_gs] = {
                    "gs": gs,
                    "rawText": "\n".join(raw_text_parts)[:8000],
                    "fields": values,
                }
            return bool(summary["ocrByGs"])
        finally:
            workbook.close()

    try:
        if read_ocr_workbook(path):
            return summary

        search_roots = []
        if output_root:
            search_roots.append(Path(output_root))
        search_roots.append(path.parent)
        candidates: list[Path] = []
        for root in search_roots:
            if root.is_dir():
                candidates.extend(root.glob("OCR결과_*.xlsx"))
        candidates = sorted(set(candidates), key=lambda item: item.stat().st_mtime, reverse=True)
        for candidate in candidates:
            if read_ocr_workbook(candidate):
                summary["linkedFrom"] = str(path)
                return summary
    except Exception as exc:
        summary["available"] = False
        summary["error"] = str(exc)
    return summary


def append_keyword_record(target: dict, base_gs: str, field: str, values: object) -> None:
    if not base_gs:
        return
    record = target.setdefault(base_gs, {
        "productNames": [],
        "searchTerms": [],
        "longtails": [],
        "debugTerms": [],
    })
    record[field] = unique_terms(record.get(field, []) + split_candidate_terms(values))


def read_pipeline_keyword_summary(output_file: str) -> dict:
    path = Path(output_file or "")
    summary = {
        "available": False,
        "workbook": str(path) if output_file else "",
        "keywordByGs": {},
    }
    if not output_file or not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return summary
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            pending: list[tuple[str, object]] = []
            for sheet_name in workbook.sheetnames:
                if sheet_name not in {"분리추출후", "B마켓", "디버그"}:
                    continue
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                headers = [text_value(value) for value in next(rows, [])]
                if not headers:
                    continue
                for row in rows:
                    values = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers)) if headers[i]}
                    gs = extract_gs(
                        values.get("자체 상품코드", ""),
                        values.get("적용 GS코드", ""),
                        values.get("상품명(원본)", ""),
                        values.get("공급사 상품명", ""),
                    )
                    base_gs = split_gs(gs)[0] if gs else ""
                    product_names = [
                        values.get("상품명", ""),
                        values.get("최종상품명", ""),
                        values.get("기본상품명(옵션제외)", ""),
                    ]
                    search_terms = [
                        values.get("검색어설정", ""),
                        values.get("검색어설정(R)", ""),
                        values.get("검색키워드", ""),
                        values.get("키워드정렬", ""),
                        values.get("로컬빌더", ""),
                    ]
                    longtails = [
                        values.get("롱테일10", ""),
                        values.get("앵커", ""),
                        values.get("베이스라인", ""),
                    ]
                    debug_terms = [
                        values.get("상세요약", ""),
                        values.get("옵션(본문)", ""),
                        values.get("옵션_토큰", ""),
                    ]
                    if base_gs:
                        append_keyword_record(summary["keywordByGs"], base_gs, "productNames", " ".join(text_value(v) for v in product_names))
                        append_keyword_record(summary["keywordByGs"], base_gs, "searchTerms", " ".join(text_value(v) for v in search_terms))
                        append_keyword_record(summary["keywordByGs"], base_gs, "longtails", " ".join(text_value(v) for v in longtails))
                        append_keyword_record(summary["keywordByGs"], base_gs, "debugTerms", " ".join(text_value(v) for v in debug_terms))
                    else:
                        for field, vals in (
                            ("productNames", product_names),
                            ("searchTerms", search_terms),
                            ("longtails", longtails),
                            ("debugTerms", debug_terms),
                        ):
                            joined = " ".join(text_value(v) for v in vals)
                            if joined:
                                pending.append((field, joined))
            if len(summary["keywordByGs"]) == 1 and pending:
                only_base = next(iter(summary["keywordByGs"]))
                for field, value in pending:
                    append_keyword_record(summary["keywordByGs"], only_base, field, value)
            summary["available"] = bool(summary["keywordByGs"])
        finally:
            workbook.close()
    except Exception as exc:
        summary["error"] = str(exc)
    return summary


def find_processed_listing_images(output_root: object, base_gs: str, min_count: int = 3) -> dict:
    base = text_value(base_gs).upper()
    if not base:
        return {"urls": [], "folder": "", "status": "missing"}
    search_roots: list[Path] = []
    root_value = text_value(output_root)
    if root_value:
        search_roots.append(Path(root_value))
    search_roots.append(EXPORT_ROOT)

    listing_roots: list[Path] = []
    for root in search_roots:
        root = root.resolve()
        candidates = [root / "listing_images"]
        if root.name.lower() == "listing_images":
            candidates.insert(0, root)
        for candidate in candidates:
            if candidate.is_dir() and candidate not in listing_roots:
                listing_roots.append(candidate)

    folders: list[Path] = []
    for listing_root in listing_roots:
        direct_candidates = [
            listing_root / base,
            *(path for path in listing_root.glob(f"*/{base}") if path.is_dir()),
        ]
        for folder in direct_candidates:
            if folder.is_dir() and folder not in folders:
                folders.append(folder)

    folders.sort(key=lambda folder: folder.stat().st_mtime, reverse=True)
    best_files: list[Path] = []
    best_folder = ""
    for folder in folders:
        files = sorted(
            [path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS],
            key=image_file_sort_key,
        )
        if len(files) > len(best_files):
            best_files = files
            best_folder = str(folder)
        if len(files) >= min_count:
            return {
                "urls": [runtime_web_url(path) for path in files],
                "folder": str(folder),
                "status": "ready",
            }

    return {
        "urls": [],
        "folder": best_folder,
        "status": "insufficient" if best_files else "missing",
        "foundCount": len(best_files),
        "requiredCount": min_count,
    }


def hydrate_seed_payload(seed_payload: dict) -> dict:
    if not isinstance(seed_payload, dict):
        return seed_payload
    output_root = (
        seed_payload.get("pipelineResult", {}).get("output_root", "")
        if isinstance(seed_payload.get("pipelineResult"), dict)
        else ""
    ) or (
        seed_payload.get("artifacts", {}).get("outputRoot", "")
        if isinstance(seed_payload.get("artifacts"), dict)
        else ""
    )
    products = seed_payload.get("products") if isinstance(seed_payload.get("products"), list) else []
    for product in products:
        if not isinstance(product, dict):
            continue
        base_gs = text_value(product.get("baseGs")) or split_gs(product.get("gs", ""))[0]
        if not isinstance(product.get("naverProvidedNotice"), dict):
            ocr = product.get("ocrAnalysis") if isinstance(product.get("ocrAnalysis"), dict) else {}
            product["naverProvidedNotice"] = build_naver_provided_notice({
                "gs": product.get("gs", ""),
                "baseGs": base_gs,
                "name": product.get("sourceName", ""),
                "opt": product.get("optionSummary", ""),
                "optionItems": product.get("optionItems", []),
            }, ocr)
        processed = find_processed_listing_images(output_root, base_gs)
        processed_images = processed.get("urls") if isinstance(processed.get("urls"), list) else []
        if not processed_images:
            continue
        images = product.get("images") if isinstance(product.get("images"), dict) else {}
        images.update({
            "sourceThumb": processed_images[0],
            "representative": processed_images[0],
            "listingCandidates": processed_images,
            "additional": processed_images[1:20],
            "selectionSource": "processed_listing_images",
            "processedFolder": processed.get("folder", ""),
            "processedStatus": processed.get("status", ""),
            "processedFoundCount": len(processed_images),
            "processedRequiredCount": 3,
        })
        product["images"] = images
    return seed_payload


def build_seed_products(
    source_path: str,
    selected_gs: list[str],
    ocr_summary: dict,
    keyword_summary: dict | None = None,
    output_root: object = "",
) -> list[dict]:
    parsed = parse_source_preview(Path(source_path), max_preview=100000)
    selected = set(selected_gs or [])
    ocr_by_gs = (ocr_summary or {}).get("ocrByGs", {})
    keyword_by_gs = (keyword_summary or {}).get("keywordByGs", {})
    products: list[dict] = []
    for row in parsed.get("preview", []):
        if selected and row.get("gs") not in selected and row.get("baseGs") not in selected:
            continue
        base_gs = row.get("baseGs") or split_gs(row.get("gs", ""))[0]
        ocr_record = ocr_by_gs.get(base_gs) or ocr_by_gs.get(row.get("gs", "")) or {}
        keyword_record = keyword_by_gs.get(base_gs) or keyword_by_gs.get(row.get("gs", "")) or {}
        ocr_text = ocr_record.get("rawText", "")
        option_summary = row.get("opt", "단일")
        option_items = row.get("optionItems", [])
        option_info = option_meta(option_summary, option_items)
        processed = find_processed_listing_images(output_root, base_gs)
        processed_images = processed.get("urls") if isinstance(processed.get("urls"), list) else []
        raw_listing_images = unique_terms(row.get("listingImages") or [])
        raw_source_thumb = row.get("thumb", "") or (raw_listing_images[0] if raw_listing_images else "")
        detail_image_set = set(row.get("detailImages") or [])
        if processed_images:
            listing_images = processed_images
            source_thumb = processed_images[0]
            additional_images = processed_images[1:20]
            selection_source = "processed_listing_images"
        else:
            listing_images = raw_listing_images
            source_thumb = raw_source_thumb
            additional_images = [
                url
                for url in unique_terms((row.get("additionalImages") or []) + [url for url in listing_images if url != source_thumb])
                if url and url != source_thumb and url not in detail_image_set
            ][:20]
            selection_source = "source_listing_columns_only"
        naver_provided_notice = build_naver_provided_notice(row, ocr_record)
        products.append({
            "gs": row.get("gs", ""),
            "baseGs": base_gs,
            "sourceName": row.get("name", ""),
            "price": row.get("price", 0),
            "supplyPrice": row.get("supplyPrice", 0),
            "salePrice": row.get("salePrice", row.get("price", 0)),
            "consumerPrice": row.get("consumerPrice", 0),
            "optionSummary": option_summary,
            "optionInput": row.get("optionInput", ""),
            "optionAdditionalAmounts": row.get("optionAdditionalAmounts", []),
            "optionType": option_info["optionType"],
            "optionCount": option_info["optionCount"],
            "optionItems": option_items,
            "detailHtml": row.get("detailHtml", ""),
            "images": {
                "sourceThumb": source_thumb,
                "representative": source_thumb,
                "listingCandidates": listing_images,
                "additional": additional_images,
                "detail": row.get("detailImages", []),
                "processedSize": "1000x1000",
                "selectionSource": selection_source,
                "processedFolder": processed.get("folder", ""),
                "processedStatus": processed.get("status", ""),
                "processedFoundCount": processed.get("foundCount", len(processed_images)),
                "processedRequiredCount": processed.get("requiredCount", 3),
            },
            "ocrAnalysis": {
                "status": "loaded" if ocr_record else "pending",
                "rawText": ocr_text,
                "fields": ocr_record.get("fields", {}),
            },
            "naverProvidedNotice": naver_provided_notice,
            "photoAnalysis": {
                "status": "pending",
                "facts": [],
                "notes": "대표/추가 이미지 분석 결과를 여기에 누적한다.",
            },
            "keywordCandidatePool": seed_keyword_pool_from_product(row, ocr_text, keyword_record),
            "generatedKeywordSeed": keyword_record,
            "reviewFields": {
                "productName": "",
                "searchTerms": "",
                "memo": "",
            },
        })
    return products


def normalize_keyword_channels(channels: list[str] | None, account_scope: str = "전체") -> list[str]:
    raw_channels = channels or [f"{account}:{market}" for account in ("A", "B") for market in SALES_MARKETS]
    out: list[str] = []
    seen: set[str] = set()
    for raw in raw_channels:
        key = text_value(raw)
        if ":" not in key:
            continue
        account, market = key.split(":", 1)
        account = account.upper()
        if account not in {"A", "B"} or market not in SALES_MARKETS:
            continue
        if account_scope and account_scope != "전체" and account != account_scope:
            continue
        normalized = f"{account}:{market}"
        if normalized in seen:
            continue
        seen.add(normalized)
        out.append(normalized)
    return out


def build_keyword_job_products(seed_payload: dict, selected_gs: list[str]) -> list[dict]:
    selected = set(selected_gs or [])
    products = seed_payload.get("products") if isinstance(seed_payload.get("products"), list) else []
    out: list[dict] = []
    for product in products:
        if not isinstance(product, dict):
            continue
        gs = text_value(product.get("gs"))
        base_gs = text_value(product.get("baseGs"))
        if selected and gs not in selected and base_gs not in selected:
            continue
        ocr = product.get("ocrAnalysis") if isinstance(product.get("ocrAnalysis"), dict) else {}
        photo = product.get("photoAnalysis") if isinstance(product.get("photoAnalysis"), dict) else {}
        out.append({
            "gs": gs,
            "baseGs": base_gs,
            "sourceName": product.get("sourceName", ""),
            "price": product.get("price", 0),
            "supplyPrice": product.get("supplyPrice", 0),
            "salePrice": product.get("salePrice", product.get("price", 0)),
            "consumerPrice": product.get("consumerPrice", 0),
            "optionSummary": product.get("optionSummary", ""),
            "optionInput": product.get("optionInput", ""),
            "optionAdditionalAmounts": product.get("optionAdditionalAmounts", []),
            "optionType": product.get("optionType", ""),
            "optionCount": product.get("optionCount", 0),
            "optionItems": product.get("optionItems", []),
            "detailHtml": product.get("detailHtml", ""),
            "images": product.get("images", {}),
            "ocrAnalysis": {
                "status": ocr.get("status", ""),
                "rawText": text_value(ocr.get("rawText"))[:5000],
                "fields": ocr.get("fields", {}),
            },
            "photoAnalysis": photo,
            "keywordCandidatePool": product.get("keywordCandidatePool", {}),
            "generatedKeywordSeed": product.get("generatedKeywordSeed", {}),
        })
    return out


def build_keyword_prompt(input_file: str, output_file: str) -> str:
    return f"""
너는 한국 오픈마켓 상품명/검색어 생성 담당자다.
현재 작업 폴더의 `{input_file}`을 읽고, 반드시 `{output_file}` 하나만 UTF-8 JSON으로 작성해라.
코드 파일이나 다른 파일은 수정하지 마라.

목표:
- Cafe24는 실제 판매 채널이 아니다. 공통 후보 풀로만 참고한다.
- 실제 생성 채널은 네이버/쿠팡/롯데ON/11번가/ESM의 A/B 계정이다.
- 입력 상품마다 요청된 모든 채널의 상품명과 검색어를 실제로 새로 작성한다.
- 단순 후보 나열이 아니라 상품 정체성, 기능, 사용처, 문제해결, 재질/규격, 현장명/동의어를 판단해서 조합한다.

출력 JSON 스키마:
{{
  "schema": "webocr.keyword.v1",
  "provider": "codex-cli",
  "products": [
    {{
      "gs": "GS코드",
      "channels": {{
        "A:네이버": {{
          "title": "상품명",
          "searchTerms": "검색어 또는 태그 문자열",
          "tags": ["태그1", "태그2"],
          "candidateCount": 12,
          "notes": "짧은 판단 근거"
        }}
      }}
    }}
  ]
}}

생성 규칙:
- 상품명은 정확하게, 검색어/태그는 넓게 간다.
- 상품명은 후보 단어를 단순히 이어붙이지 말고 아래 공식으로 만든다.
  1) 대표 상품 정체성 또는 표준명
  2) 실제 기능/효과
  3) 주 사용처 또는 구매 상황
  4) 재질/규격은 상품 식별에 중요할 때만 추가
- 상품명이 짧아질 때는 수량을 넣지 말고 기능/사용처/문제해결 단어로 채운다.
- 좋은 상품명 예시 구조: `카라비너 릴고리 와이어 릴홀더 백팩 스트랩 연결고리`, `쿠션 깔창 신발 밑창 보강 패드 운동화 구두 PU`.
- 네이버 상품명은 35-48자 사이를 우선한다. 50자를 넘기지 않는다.
- 네이버는 SEO 기준을 우선한다. 상품명 공식은 `브랜드/제조사 + 모델명/모델코드 + 상품유형/카테고리 + 색상/용도/주요속성`이지만, 브랜드/모델이 없으면 빼고 핵심 카테고리와 용도/속성을 앞에 둔다.
- 네이버에서 참고할 항목은 브랜드/제조사, 시리즈, 모델명/모델코드, 상품 유형/카테고리, 색상, 소재, 구성품/수량, 사이즈, 대상 성별/연령, 용량/규격/주요 속성, 판매 옵션이다.
- 네이버 상품명은 중요한 키워드를 앞쪽에 배치하고, 중복 단어/관련 없는 키워드/할인/세일/무료배송/광고 문구를 제외한다.
- 네이버는 한글 중심으로 작성하고 필요한 경우에만 영문/숫자를 사용한다.
- 쿠팡 상품명은 네이버보다 길게 가능하되 60-80자 안에서 자연스럽게 만든다.
- 롯데ON은 네이버와 쿠팡 사이의 중간형으로 만든다.
- 11번가와 ESM은 엑셀 업로드용이지만 너무 짧게 만들지 않는다. 45-62자 안에서 상품 유형, 기능, 사용처, 재질/규격, 문제해결어를 더 보강한다.
- 11번가/ESM 검색어는 최소 14개 이상을 목표로 하고, 표준어 + 현장명 + 동의어 + 사용처 + 문제해결어를 충분히 넣는다.
- A계정은 표준어/대표어 중심, B계정은 동의어/사용처/현장명 중심으로 다르게 만든다.
- A/B 계정은 같은 단어만 순서 바꾸지 말고, 대표어와 현장명을 실제로 다르게 섞는다.
- 색상 옵션은 상품명에 넣지 않는다. 옵션 컬럼에서 처리한다.
- 옵션형 사이즈는 전체 나열하지 말고 필요하면 `230-270 사이즈 선택형`처럼 압축한다.
- `1개`, `1입`, `1매`, `1p`처럼 단품을 뜻하는 수량은 상품명/검색어에 넣지 않는다. 노출 가치가 낮다.
- `2개`, `5매`, `10개`, `100p`, `3세트`처럼 실제 구성 차이를 만드는 수량만 검색어에 남길 수 있다.
- OCR에서만 나온 숫자, 눈금, 이미지 배경 숫자는 상품명/검색어에서 제외한다.
- 무료배송/할인/추천/인기/베스트/가격/문의/상담/판매자명 같은 판매 문구는 금지한다.
- `발편한`, `발 편한`, `편한발`처럼 감성/광고형 문구는 상품명과 검색어에 넣지 않는다. `쿠션`, `충격 완화`, `착용감 보강`처럼 기능어로 바꾼다.
- 같은 단어를 반복하지 않는다.
- 상품명에서는 붙여쓴 합성어를 자연스럽게 띄어쓴다. 예: 카라비너릴고리 -> 카라비너 릴고리, 와이어릴고리 -> 와이어 릴고리, 쿠션깔창 -> 쿠션 깔창, 보강패드 -> 보강 패드, 충격완화 -> 충격 완화.
- 검색어/태그는 12-20개를 목표로 하며, 상품명에 못 넣은 동의어/현장명/사용처/문제해결어를 우선 보강한다.
- 검색어에는 쉼표로 구분된 실제 검색 가능한 말만 넣고, 무관한 인기어와 일부러 만든 오타는 넣지 않는다.
- 검색어/태그의 각 항목은 공백 없이 붙여쓴다. 예: `쿠션 깔창`은 검색어에서 `쿠션깔창`, `보강 패드`는 `보강패드`, `충격 완화`는 `충격완화`로 쓴다.
- `title`, `searchTerms`, `tags`는 비워두지 않는다.
- 요청 채널이 10개면 상품마다 10개 채널 모두 작성한다.

입력 JSON에는 `channels`, `products`, `policy`가 들어 있다.
최종 응답은 설명하지 말고 파일 작성만 끝내라.
""".strip()


SINGLE_QUANTITY_RE = re.compile(
    r"(?<![0-9A-Za-z가-힣])1\s*(?:개|입|매|p|P|pcs|PCS|pc|PC)(?![0-9A-Za-z가-힣])"
)


def strip_low_value_single_quantity(value: str) -> str:
    cleaned = SINGLE_QUANTITY_RE.sub(" ", text_value(value))
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    cleaned = re.sub(r"\s*,\s*", ", ", cleaned)
    cleaned = re.sub(r"(?:^|,\s*)$", "", cleaned)
    return cleaned.strip(" ,/")


def split_search_keyword_chunks(value: object) -> list[str]:
    raw = text_value(value)
    if not raw:
        return []
    raw = re.sub(r"<[^>]+>", " ", raw)
    chunks: list[str] = []
    for chunk in re.split(r"[,/|;·\n\r\t]+", raw):
        text = normalize_term(chunk)
        if text:
            chunks.append(text)
    return unique_terms(chunks)


def clean_keyword_terms(value: str) -> str:
    terms = [
        compact_search_keyword(strip_low_value_single_quantity(term))
        for term in split_search_keyword_chunks(value)
    ]
    return ", ".join(unique_terms([term for term in terms if term]))


def validate_keyword_result(payload: dict, products: list[dict], channels: list[str]) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("keyword result is not an object")
    raw_products = payload.get("products")
    if not isinstance(raw_products, list):
        raise ValueError("keyword result has no products list")

    product_keys = {text_value(product.get("gs")) or text_value(product.get("baseGs")) for product in products}
    normalized_products: list[dict] = []
    for item in raw_products:
        if not isinstance(item, dict):
            continue
        gs = text_value(item.get("gs"))
        if gs not in product_keys:
            continue
        raw_channels = item.get("channels") if isinstance(item.get("channels"), dict) else {}
        channel_map: dict[str, dict] = {}
        for channel in channels:
            value = raw_channels.get(channel)
            if not isinstance(value, dict):
                continue
            title = clean_product_title(strip_low_value_single_quantity(value.get("title")))
            search_terms = clean_keyword_terms(value.get("searchTerms") or value.get("search_terms"))
            tags = value.get("tags") if isinstance(value.get("tags"), list) else split_candidate_terms(search_terms)
            tags = unique_terms([
                compact_search_keyword(strip_low_value_single_quantity(tag))
                for tag in tags
                if compact_search_keyword(strip_low_value_single_quantity(tag))
            ])[:30]
            if not title or not search_terms:
                continue
            candidate_count = len(tags) or len(split_candidate_terms(search_terms))
            channel_map[channel] = {
                "title": title,
                "searchTerms": search_terms,
                "tags": tags,
                "candidateCount": candidate_count,
                "notes": text_value(value.get("notes"))[:500],
                "provider": text_value(payload.get("provider")) or "codex-cli",
                "generatedAt": now_text(),
            }
        if channel_map:
            normalized_products.append({"gs": gs, "channels": channel_map})
    if not normalized_products:
        raise ValueError("keyword result has no usable channel data")
    return {
        "schema": "webocr.keyword.v1",
        "provider": text_value(payload.get("provider")) or "codex-cli",
        "products": normalized_products,
    }


def apply_keyword_result_to_seed(seed_payload: dict, keyword_result: dict, channels: list[str]) -> None:
    by_gs = {
        text_value(item.get("gs")): item.get("channels", {})
        for item in keyword_result.get("products", [])
        if isinstance(item, dict)
    }
    for product in seed_payload.get("products", []):
        if not isinstance(product, dict):
            continue
        gs = text_value(product.get("gs"))
        channel_map = by_gs.get(gs)
        if not channel_map:
            continue
        existing = product.get("marketKeywords") if isinstance(product.get("marketKeywords"), dict) else {}
        for channel in channels:
            if channel in channel_map:
                current_value = existing.get(channel) if isinstance(existing.get(channel), dict) else {}
                next_value = channel_map[channel]
                current_title = text_value(current_value.get("title"))
                next_title = text_value(next_value.get("title"))
                current_count = int(current_value.get("candidateCount") or len(current_value.get("tags") or []))
                next_count = int(next_value.get("candidateCount") or len(next_value.get("tags") or []))
                current_score = len(current_title) + current_count * 4
                next_score = len(next_title) + next_count * 4
                if current_title and current_score > next_score + 18:
                    current_value["keptBecause"] = "regeneration_result_was_weaker"
                    existing[channel] = current_value
                else:
                    existing[channel] = next_value
        product["marketKeywords"] = existing
    seed_payload["keywordGeneration"] = {
        "status": "completed",
        "provider": keyword_result.get("provider", "codex-cli"),
        "channels": channels,
        "updatedAt": now_text(),
        "note": "키워드 생성 단계에서 Codex CLI를 호출해 실제 마켓별 상품명/검색어를 생성했다.",
    }


def read_csv_rows(path: Path) -> list[dict]:
    for encoding in ("utf-8-sig", "cp949", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                return list(csv.DictReader(handle, dialect=dialect))
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle))


def read_excel_rows(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [text_value(value) for value in next(rows, [])]
        records = []
        for row in rows:
            records.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers)) if headers[i]})
        workbook.close()
        return records

    import pandas as pd

    frame = pd.read_excel(path, dtype=object)
    frame = frame.where(pd.notnull(frame), "")
    return frame.to_dict(orient="records")


def read_source_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xls"}:
        return read_excel_rows(path)
    return []


def collect_values(row: dict, headers: dict[str, str], names: list[str]) -> list[str]:
    values: list[str] = []
    for name in names:
        key = headers.get(normalize_header(name))
        if key is not None:
            value = text_value(row.get(key, ""))
            if value:
                values.append(value)
    return values


def extract_image_urls(value: str) -> list[str]:
    text = text_value(value)
    if not text:
        return []
    text = text.replace("&amp;", "&")
    urls = re.findall(r"<img[^>]+src=[\"']?([^\"'>\s]+)", text, flags=re.IGNORECASE)
    urls.extend(re.findall(r"https?://[^\s\"'<>|,;]+", text, flags=re.IGNORECASE))
    urls.extend(re.findall(r"//[^\s\"'<>|,;]+", text, flags=re.IGNORECASE))
    urls.extend(re.findall(r"(?:[A-Za-z]:\\|/)[^\s\"'<>|,;]+?\.(?:jpg|jpeg|png|webp|bmp)", text, flags=re.IGNORECASE))
    for part in re.split(r"[\n\r\t,;|]+", text):
        clean_part = part.strip().strip("\"'")
        if re.search(r"\.(?:jpg|jpeg|png|webp|bmp)(?:\?.*)?$", clean_part, re.IGNORECASE):
            urls.append(clean_part)
    out: list[str] = []
    seen: set[str] = set()
    for url in urls:
        clean = urllib.parse.unquote(text_value(url).strip().strip("\"'"))
        if clean.startswith("//"):
            clean = f"https:{clean}"
        if not clean or clean in seen:
            continue
        seen.add(clean)
        out.append(clean)
    return out


def collect_image_urls(row: dict, headers: dict[str, str], names: list[str]) -> list[str]:
    return unique_image_urls([
        url
        for value in collect_values(row, headers, names)
        for url in extract_image_urls(value)
    ])


def image_url_identity(value: object) -> str:
    raw = text_value(value)
    if not raw:
        return ""
    parsed = urllib.parse.urlparse(raw)
    if parsed.scheme in {"http", "https"}:
        host = parsed.netloc.lower()
        path = urllib.parse.unquote(parsed.path).lower()
        return f"{host}{path}"
    return raw.replace("\\", "/").lower()


def unique_image_urls(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = text_value(value)
        key = image_url_identity(clean)
        if not clean or not key or key in seen:
            continue
        seen.add(key)
        out.append(clean)
    return out


def parse_source_preview(path: Path, max_preview: int = 500) -> dict:
    records = read_source_records(path)
    if not records:
        return {"rows": 0, "gsCodes": 0, "preview": [], "columns": []}

    columns = list(records[0].keys()) if records else []
    headers = {normalize_header(col): col for col in columns}
    groups: dict[str, list[dict]] = {}
    raw_gs_codes: set[str] = set()
    for index, row in enumerate(records):
        code = first_value(row, headers, ["자체 상품코드", "GS코드", "gs코드", "상품코드", "판매자상품코드"])
        name_raw = first_value(row, headers, ["약식 상품명", "약식상품명", "상품명(관리용)", "상품명", "공급사 상품명", "상품 요약설명"])
        gs = extract_gs(code, name_raw, row)
        if not gs:
            continue
        base_gs, suffix = split_gs(gs)
        raw_gs_codes.add(gs)
        price_info = source_price_info(row, headers)
        opt = first_value(row, headers, ["옵션입력", "옵션", "옵션세트명", "옵션명", "규격"]) or "단일"
        listing_images = collect_image_urls(row, headers, LISTING_IMAGE_COLUMNS)
        additional_images = collect_image_urls(row, headers, ADDITIONAL_IMAGE_COLUMNS)
        detail_images = collect_image_urls(row, headers, DETAIL_IMAGE_COLUMNS)
        detail_html = first_value(row, headers, DETAIL_HTML_COLUMNS)
        thumb = listing_images[0] if listing_images else ""

        groups.setdefault(base_gs, []).append({
            "index": index,
            "gs": gs,
            "baseGs": base_gs,
            "suffix": suffix,
            "nameRaw": name_raw,
            "name": short_name(name_raw, gs),
            "price": price_info["salePrice"],
            "supplyPrice": price_info["supplyPrice"],
            "salePrice": price_info["salePrice"],
            "consumerPrice": price_info["consumerPrice"],
            "opt": opt,
            "thumb": thumb,
            "listingImages": listing_images,
            "additionalImages": additional_images,
            "detailImages": detail_images,
            "detailHtml": detail_html,
        })

    preview = []
    ordered_groups = sorted(groups.items(), key=lambda pair: min(item["index"] for item in pair[1]))
    for base_gs, items in ordered_groups[:max_preview]:
        items = sorted(items, key=lambda item: item["index"])
        representative = next((item for item in items if item["suffix"] == "A"), items[0])
        clean_names = [item["name"] for item in items if item["name"]]
        base_name = common_word_prefix(clean_names) if len(items) > 1 else ""
        if not base_name:
            base_name = representative["name"] or representative["gs"]

        option_items = []
        option_labels = []
        for item in items:
            label = strip_option_label(item["name"], base_name)
            if not label and item["opt"] and item["opt"] != "단일":
                label = item["opt"]
            if not label and len(items) > 1:
                label = item["suffix"] or item["gs"].replace(base_gs, "", 1) or item["name"]
            if label:
                option_labels.append(label)
            option_items.append({
                "gs": item["gs"],
                "suffix": item["suffix"],
                "name": item["name"],
                "option": label or "단일",
                "price": item["price"],
                "supplyPrice": item.get("supplyPrice", 0),
                "salePrice": item.get("salePrice", item["price"]),
                "consumerPrice": item.get("consumerPrice", 0),
                "thumb": item["thumb"],
            })

        listing_images = unique_image_urls([
            url
            for item in items
            for url in (item.get("listingImages", []) + ([item.get("thumb", "")] if item.get("thumb") else []))
        ])
        detail_images = unique_image_urls([
            url
            for item in items
            for url in item.get("detailImages", [])
        ])
        detail_html = next((text_value(item.get("detailHtml")) for item in items if text_value(item.get("detailHtml"))), "")
        representative_thumb = representative["thumb"] or (listing_images[0] if listing_images else "")
        additional_images = unique_image_urls([
            url
            for item in items
            for url in item.get("additionalImages", [])
        ] + [url for url in listing_images if url != representative_thumb])
        detail_keys = {image_url_identity(url) for url in detail_images}
        representative_key = image_url_identity(representative_thumb)
        additional_images = [
            url for url in additional_images
            if image_url_identity(url) not in detail_keys and image_url_identity(url) != representative_key
        ][:20]
        base_sale = option_items[0]["salePrice"] if option_items else representative["price"]
        option_additionals = [
            int((item.get("salePrice") or item.get("price") or 0) - base_sale)
            for item in option_items
        ] if len(option_items) > 1 else []
        option_input = "옵션{" + "|".join(
            f"{chr(65 + index)} {text_value(item.get('option'))}"
            for index, item in enumerate(option_items[:26])
            if text_value(item.get("option")) and text_value(item.get("option")) != "단일"
        ) + "}" if len(option_items) > 1 else ""

        preview.append({
            "id": f"row-{len(preview) + 1}",
            "gs": representative["gs"],
            "baseGs": base_gs,
            "name": base_name,
            "price": representative["price"],
            "supplyPrice": representative.get("supplyPrice", 0),
            "salePrice": representative.get("salePrice", representative["price"]),
            "consumerPrice": representative.get("consumerPrice", 0),
            "opt": compact_option_summary(option_labels, len(items)),
            "optionInput": option_input,
            "optionAdditionalAmounts": option_additionals,
            "thumb": representative_thumb,
            "listingImages": listing_images,
            "detailImages": detail_images,
            "detailHtml": detail_html,
            "additionalImages": additional_images,
            "optionItems": option_items,
        })
        preview[-1].update(option_meta(preview[-1]["opt"], option_items))

    return {
        "rows": len(records),
        "gsCodes": len(groups),
        "rawGsCodes": len(raw_gs_codes),
        "preview": preview,
        "columns": columns,
    }


def create_selected_source_file(source_path: str, selected_gs: list[str]) -> tuple[str, dict]:
    source = Path(source_path)
    selected_codes = {text_value(value).upper() for value in selected_gs or [] if text_value(value)}
    selected_bases = {split_gs(code)[0] for code in selected_codes if code}
    if not selected_bases:
        raise ValueError("selectedGs is empty")

    records = read_source_records(source)
    if not records:
        raise ValueError("source file has no rows")

    columns = list(records[0].keys())
    headers = {normalize_header(col): col for col in columns}
    filtered: list[dict] = []
    matched_codes: set[str] = set()
    for row in records:
        code = first_value(row, headers, ["자체 상품코드", "GS코드", "gs코드", "상품코드", "판매자상품코드"])
        name_raw = first_value(row, headers, ["약식 상품명", "약식상품명", "상품명(관리용)", "상품명", "공급사 상품명", "상품 요약설명"])
        gs = extract_gs(code, name_raw, row)
        if not gs:
            continue
        base_gs, _ = split_gs(gs)
        if gs in selected_codes or base_gs in selected_bases:
            filtered.append(row)
            matched_codes.add(gs)

    if not filtered:
        raise ValueError(f"selected GS not found in source: {', '.join(sorted(selected_codes))}")

    for row in filtered:
        for key in row.keys():
            if key not in columns:
                columns.append(key)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = UPLOAD_ROOT / safe_name(f"{source.stem}_{stamp}_selected_{len(selected_bases)}.csv")
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(filtered)

    return str(target), {
        "sourceFile": str(source),
        "filteredFile": str(target),
        "selectedProducts": len(selected_bases),
        "rawRows": len(records),
        "filteredRows": len(filtered),
        "matchedGs": sorted(matched_codes),
    }


def parse_multipart(body: bytes, content_type: str) -> dict[str, dict]:
    match = re.search(r"boundary=(?P<boundary>[^;]+)", content_type or "")
    if not match:
        raise ValueError("multipart boundary not found")
    boundary = match.group("boundary").strip('"').encode("utf-8")
    out: dict[str, dict] = {}
    for raw in body.split(b"--" + boundary):
        raw = raw.strip()
        if not raw or raw == b"--":
            continue
        if raw.endswith(b"--"):
            raw = raw[:-2].strip()
        if b"\r\n\r\n" not in raw:
            continue
        header_blob, content = raw.split(b"\r\n\r\n", 1)
        if content.endswith(b"\r\n"):
            content = content[:-2]
        headers = header_blob.decode("utf-8", errors="replace")
        disposition = next((line for line in headers.split("\r\n") if line.lower().startswith("content-disposition:")), "")
        name_match = re.search(r'name="([^"]+)"', disposition)
        if not name_match:
            continue
        filename_match = re.search(r'filename="([^"]*)"', disposition)
        field_name = name_match.group(1)
        out[field_name] = {
            "filename": filename_match.group(1) if filename_match else "",
            "content": content,
            "headers": headers,
        }
    return out


def py_command() -> list[str]:
    py = shutil.which("py")
    if py:
        return [py, "-3"]
    return [sys.executable]


def bool_arg(value: object) -> str:
    return "true" if bool(value) else "false"


def bridge_args(source_path: str, settings: dict) -> list[str]:
    phase = "full"
    return [
        *py_command(),
        str(BRIDGE_SCRIPT),
        "--legacy-root", str(BACKEND_ROOT),
        "--source", source_path,
        "--make-listing", "true",
        "--listing-size", "1000",
        "--listing-pad", str(settings.get("ListingPad", 20)),
        "--listing-max", str(settings.get("ListingMax", 20)),
        "--logo-path", str(settings.get("LogoPath", "") or ""),
        "--logo-ratio", str(settings.get("LogoRatio", 14)),
        "--logo-opacity", str(settings.get("LogoOpacity", 65)),
        "--logo-pos", str(settings.get("LogoPosition", "tr") or "tr"),
        "--use-auto-contrast", bool_arg(settings.get("UseAutoContrast", True)),
        "--use-sharpen", bool_arg(settings.get("UseSharpen", True)),
        "--use-small-rotate", bool_arg(settings.get("UseSmallRotate", True)),
        "--rotate-zoom", str(settings.get("RotateZoom", 1.04)),
        "--ultra-angle-deg", str(settings.get("UltraAngleDeg", 0.35)),
        "--ultra-translate-px", str(settings.get("UltraTranslatePx", 0.6)),
        "--ultra-scale-pct", str(settings.get("UltraScalePct", 0.25)),
        "--trim-tol", str(settings.get("TrimTolerance", 8)),
        "--jpeg-q-min", str(settings.get("JpegQualityMin", 88)),
        "--jpeg-q-max", str(settings.get("JpegQualityMax", 92)),
        "--flip-lr", bool_arg(settings.get("FlipLeftRight", False)),
        "--logo-path-b", str(settings.get("LogoPathB", "") or ""),
        "--img-tag", str(settings.get("ImgTag", "") or ""),
        "--img-tag-b", str(settings.get("ImgTagB", "") or ""),
        "--a-name-min", str(settings.get("ANameMin", 80)),
        "--a-name-max", str(settings.get("ANameMax", 100)),
        "--b-name-min", str(settings.get("BNameMin", 63)),
        "--b-name-max", str(settings.get("BNameMax", 98)),
        "--a-tag-count", str(settings.get("ATagCount", 20)),
        "--b-tag-count", str(settings.get("BTagCount", 14)),
        "--phase", phase,
        "--export-root", str(EXPORT_ROOT),
        "--model", "claude-sonnet-4-6",
        "--chunk-size", "10",
        "--keyword-version", "3.0",
    ]


def stop_process_tree(process: subprocess.Popen) -> None:
    if process.poll() is not None:
        return
    try:
        if os.name == "nt":
            subprocess.run(
                ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
        else:
            process.terminate()
    except Exception:
        try:
            process.kill()
        except Exception:
            pass


def register_active_process(job_id: str, process: subprocess.Popen) -> None:
    with ACTIVE_PROCESS_LOCK:
        ACTIVE_PROCESSES[job_id] = process


def unregister_active_process(job_id: str, process: subprocess.Popen | None = None) -> None:
    with ACTIVE_PROCESS_LOCK:
        if process is None or ACTIVE_PROCESSES.get(job_id) is process:
            ACTIVE_PROCESSES.pop(job_id, None)


def stop_active_job(job_id: str) -> bool:
    with ACTIVE_PROCESS_LOCK:
        CANCELLED_JOB_IDS.add(job_id)
        process = ACTIVE_PROCESSES.get(job_id)
    if not process:
        return False
    stop_process_tree(process)
    unregister_active_process(job_id, process)
    return True


def run_seed_job(job_id: str, payload: dict) -> None:
    job_path = JOBS_ROOT / f"{job_id}.json"
    log_path = JOBS_ROOT / f"{job_id}.log"
    job = read_json(job_path, {"jobId": job_id})
    job.update({
        "status": "running",
        "startedAt": now_text(),
        "logPath": str(log_path),
        "progressPercent": 2,
        "currentStage": "작업 준비",
    })
    write_json(job_path, job)

    source_path = payload.get("sourcePath") or payload.get("sourceFilePath") or ""
    selected_gs = payload.get("selectedGs", [])
    settings = payload.get("listingImageSettings") or {}
    if not source_path or not Path(source_path).exists():
        job.update({"status": "failed", "finishedAt": now_text(), "error": "source file not found"})
        write_json(job_path, job)
        return

    try:
        effective_source_path, filter_meta = create_selected_source_file(source_path, selected_gs)
    except Exception as exc:
        job.update({"status": "failed", "finishedAt": now_text(), "error": str(exc)})
        write_json(job_path, job)
        return

    job.update({
        "selectedGs": selected_gs,
        "filter": filter_meta,
        "progressPercent": 8,
        "currentStage": "선택 상품 원본 필터링",
        "updatedAt": now_text(),
    })
    write_json(job_path, job)

    cmd = bridge_args(effective_source_path, settings)
    result_payload: dict | None = None
    try:
        with log_path.open("w", encoding="utf-8", errors="replace") as log:
            log.write(f"[{now_text()}] START sourceToSeed\n")
            log.write(f"selected products: {filter_meta['selectedProducts']} / filtered rows: {filter_meta['filteredRows']}\n")
            log.write(f"filtered source: {effective_source_path}\n")
            log.write(" ".join(f'"{part}"' if " " in part else part for part in cmd) + "\n\n")
            process = subprocess.Popen(
                cmd,
                cwd=str(PROJECT_ROOT),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
            )
            register_active_process(job_id, process)
            assert process.stdout is not None
            tail: list[str] = []
            max_progress = 12
            result_seen = False
            try:
                for line in process.stdout:
                    clean = line.rstrip("\n")
                    log.write(clean + "\n")
                    log.flush()
                    tail = (tail + [clean])[-20:]
                    if clean.startswith("__RESULT__"):
                        result_payload = json.loads(clean[len("__RESULT__"):])
                        result_seen = True
                    line_progress, line_stage, current_gs = infer_seed_progress(clean, max_progress)
                    max_progress = max(max_progress, line_progress)
                    current = read_json(job_path, {"jobId": job_id})
                    current.update({
                        "status": "running",
                        "updatedAt": now_text(),
                        "tail": tail,
                        "progressPercent": max_progress,
                        "currentStage": line_stage,
                        "currentGs": current_gs or current.get("currentGs", ""),
                    })
                    write_json(job_path, current)
                    if result_seen:
                        break
                if result_payload:
                    try:
                        exit_code = process.wait(timeout=3)
                    except subprocess.TimeoutExpired:
                        stop_process_tree(process)
                        exit_code = 0
                else:
                    exit_code = process.wait()
            finally:
                unregister_active_process(job_id, process)

        if exit_code != 0:
            raise RuntimeError(f"pipeline failed with exit code {exit_code}")

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        source_stem = Path(source_path).stem
        seed_name = safe_name(f"{source_stem}_{stamp}.webseed.json")
        seed_path = SEED_ROOT / seed_name
        ocr_summary = read_pipeline_ocr_summary(
            (result_payload or {}).get("output_file", ""),
            (result_payload or {}).get("output_root", ""),
        )
        keyword_summary = read_pipeline_keyword_summary((result_payload or {}).get("output_file", ""))
        products = build_seed_products(
            effective_source_path,
            selected_gs,
            ocr_summary,
            keyword_summary,
            (result_payload or {}).get("output_root", ""),
        )
        job.update({"progressPercent": 92, "currentStage": "시드 파일 작성", "updatedAt": now_text()})
        write_json(job_path, job)
        seed_payload = {
            "schema": "webocr.seed.v2",
            "seedType": "source_to_seed",
            "createdAt": now_text(),
            "sourceFile": source_path,
            "effectiveSourceFile": effective_source_path,
            "sourceFilter": filter_meta,
            "selectedGs": selected_gs,
            "analysisPolicy": SEED_ANALYSIS_POLICY,
            "products": products,
            "summary": {
                "products": len(products),
                "ocrLoaded": sum(1 for product in products if product.get("ocrAnalysis", {}).get("status") == "loaded"),
                "imageRule": "1000x1000",
            },
            "pipelineResult": result_payload or {},
            "listingImageSettings": settings,
            "artifacts": {
                "outputRoot": (result_payload or {}).get("output_root", ""),
                "outputFile": (result_payload or {}).get("output_file", ""),
                "logFile": str(log_path),
            },
            "workbookSummary": {
                "sheets": ocr_summary.get("sheets", []),
                "ocrAvailable": ocr_summary.get("available", False),
                "ocrError": ocr_summary.get("error", ""),
                "keywordAvailable": keyword_summary.get("available", False),
                "keywordError": keyword_summary.get("error", ""),
            },
            "note": "1차 시드: 원본 정리, 이미지 1000x1000 가공 기준, OCR/사진분석, 키워드 후보 풀을 담는 기준 데이터셋",
        }
        write_json(seed_path, seed_payload)

        job.update({
            "status": "completed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "1차 시드 생성 완료",
            "result": {
                "seedFileName": seed_name,
                "seedPath": str(seed_path),
                "seedSize": seed_path.stat().st_size,
                "filter": filter_meta,
                "pipelineResult": result_payload or {},
            },
        })
        write_json(job_path, job)
    except Exception as exc:
        if job_id in CANCELLED_JOB_IDS:
            return
        with log_path.open("a", encoding="utf-8", errors="replace") as log:
            log.write(f"\n[{now_text()}] ERROR {exc}\n")
        job.update({"status": "failed", "finishedAt": now_text(), "error": str(exc), "currentStage": "실패"})
        write_json(job_path, job)


def run_keyword_job(job_id: str, payload: dict) -> None:
    job_path = JOBS_ROOT / f"{job_id}.json"
    log_path = JOBS_ROOT / f"{job_id}.log"
    work_dir = JOBS_ROOT / f"{job_id}_keyword"
    work_dir.mkdir(parents=True, exist_ok=True)
    job = read_json(job_path, {"jobId": job_id})
    job.update({
        "status": "running",
        "startedAt": now_text(),
        "logPath": str(log_path),
        "workDir": str(work_dir),
        "progressPercent": 3,
        "currentStage": "키워드 생성 준비",
    })
    write_json(job_path, job)

    try:
        seed_path = resolve_seed_path(payload.get("seedPath") or payload.get("sourcePath") or payload.get("path") or "")
        if not seed_path.exists():
            raise FileNotFoundError("seed file not found")
        seed_payload = read_json(seed_path, {})
        channels = normalize_keyword_channels(payload.get("channels"), text_value(payload.get("accountScope") or "전체"))
        selected_gs = payload.get("selectedGs") if isinstance(payload.get("selectedGs"), list) else []
        products = build_keyword_job_products(seed_payload, selected_gs)
        if not products:
            raise ValueError("selected seed products not found")
        if not channels:
            raise ValueError("selected market channels not found")

        input_name = "keyword_input.json"
        output_name = "keyword_result.json"
        input_payload = {
            "schema": "webocr.keyword.input.v1",
            "createdAt": now_text(),
            "seedFile": str(seed_path),
            "channels": channels,
            "options": payload.get("options", {}),
            "policy": SEED_ANALYSIS_POLICY,
            "products": products,
        }
        write_json(work_dir / input_name, input_payload)
        (work_dir / "prompt.md").write_text(build_keyword_prompt(input_name, output_name), encoding="utf-8")
        job.update({
            "seedPath": str(seed_path),
            "channels": channels,
            "selectedGs": [product["gs"] for product in products],
            "totalProducts": len(products),
            "totalChannels": len(channels),
            "progressPercent": 10,
            "currentStage": f"Codex 입력 작성 완료 · 상품 {len(products)}개 · 채널 {len(channels)}개",
            "updatedAt": now_text(),
        })
        write_json(job_path, job)

        codex_bin = shutil.which("codex.cmd") or shutil.which("codex") or "codex"
        instruction = (
            f"`prompt.md` 지시서를 먼저 읽고 `{input_name}`의 상품/채널 데이터를 분석해서 "
            f"`{output_name}`을 스키마에 맞는 JSON으로 작성해. 다른 파일은 수정하지 마."
        )
        cmd = [
            codex_bin,
            "exec",
            "--skip-git-repo-check",
            "--dangerously-bypass-approvals-and-sandbox",
            "-C",
            str(work_dir),
            instruction,
        ]
        tail: list[str] = []
        output_lines = 0
        stdout_done = object()

        def read_process_stdout(stream, output_queue: queue.Queue) -> None:
            try:
                for stdout_line in stream:
                    output_queue.put(stdout_line)
            finally:
                output_queue.put(stdout_done)

        with log_path.open("w", encoding="utf-8", errors="replace") as log:
            log.write(f"[{now_text()}] START keywordGenerate\n")
            log.write(f"seed: {seed_path}\n")
            log.write(f"products: {len(products)} / channels: {len(channels)}\n")
            log.write(" ".join(f'"{part}"' if " " in part else part for part in cmd) + "\n\n")
            process = subprocess.Popen(
                cmd,
                cwd=str(work_dir),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
            )
            register_active_process(job_id, process)
            assert process.stdout is not None
            output_queue: queue.Queue = queue.Queue()
            reader = threading.Thread(target=read_process_stdout, args=(process.stdout, output_queue), daemon=True)
            reader.start()
            reader_done = False
            started_at = time.time()
            last_heartbeat = 0.0

            def update_running_stage(stage: str) -> None:
                elapsed = int(time.time() - started_at)
                progress = min(88, max(12 + elapsed // 5, 18 + min(output_lines, 60)))
                current = read_json(job_path, {"jobId": job_id})
                current.update({
                    "status": "running",
                    "updatedAt": now_text(),
                    "tail": tail,
                    "progressPercent": progress,
                    "currentStage": stage,
                })
                write_json(job_path, current)

            try:
                while True:
                    try:
                        item = output_queue.get(timeout=0.5)
                    except queue.Empty:
                        item = None

                    if item is stdout_done:
                        reader_done = True
                    elif item is not None:
                        clean = str(item).rstrip("\n")
                        log.write(clean + "\n")
                        log.flush()
                        output_lines += 1
                        tail = (tail + [clean])[-20:]
                        update_running_stage("Codex AI 상품명/검색어 생성 중")

                    now_ts = time.time()
                    if now_ts - last_heartbeat >= 1.0:
                        stage_text = "Codex AI 응답 대기"
                        if output_lines:
                            stage_text = "Codex AI 상품명/검색어 생성 중"
                        if process.poll() is not None:
                            stage_text = "Codex 결과 파일 확인 중"
                        update_running_stage(stage_text)
                        last_heartbeat = now_ts

                    if process.poll() is not None and reader_done and output_queue.empty():
                        break
                exit_code = process.wait()
            finally:
                unregister_active_process(job_id, process)
        if exit_code != 0:
            raise RuntimeError(f"codex failed with exit code {exit_code}")

        result_path = work_dir / output_name
        if not result_path.exists():
            raise FileNotFoundError(f"{output_name} not written")
        raw_result = read_json(result_path, {})
        keyword_result = validate_keyword_result(raw_result, products, channels)
        job = read_json(job_path, job)
        job.update({"progressPercent": 90, "currentStage": "결과 검증 및 시드 반영", "updatedAt": now_text()})
        write_json(job_path, job)
        apply_keyword_result_to_seed(seed_payload, keyword_result, channels)
        write_json(seed_path, seed_payload)

        generated_products = len(keyword_result.get("products", []))
        generated_channels = sum(len(item.get("channels", {})) for item in keyword_result.get("products", []))
        job = read_json(job_path, job)
        job.update({
            "status": "completed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "키워드 생성 완료",
            "tail": tail[-20:],
            "result": {
                "seedPath": str(seed_path),
                "keywordResultPath": str(result_path),
                "products": generated_products,
                "channels": channels,
                "generatedChannels": generated_channels,
            },
        })
        write_json(job_path, job)
    except Exception as exc:
        if job_id in CANCELLED_JOB_IDS:
            return
        with log_path.open("a", encoding="utf-8", errors="replace") as log:
            log.write(f"\n[{now_text()}] ERROR {exc}\n")
        job.update({
            "status": "failed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "키워드 생성 실패",
            "error": str(exc),
        })
        write_json(job_path, job)


def normalize_upload_entries(payload: dict) -> list[dict]:
    rows = payload.get("rows")
    if not isinstance(rows, list):
        rows = []
    entries: list[dict] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        account = text_value(row.get("account", "")).upper()
        market = text_value(row.get("market", ""))
        gs = text_value(row.get("gs", "")).upper()
        if not account or not market or not gs:
            continue
        channel = text_value(row.get("channel") or row.get("channelKey") or f"{account}:{market}")
        additional_image_srcs = row.get("additionalImageSrcs")
        if not isinstance(additional_image_srcs, list):
            additional_image_srcs = extract_image_urls(row.get("additionalImageSrcs", ""))
        detail_image_srcs = row.get("detailImageSrcs")
        if not isinstance(detail_image_srcs, list):
            detail_image_srcs = extract_image_urls(row.get("detailImageSrcs", ""))
        option_items = row.get("optionItems")
        if not isinstance(option_items, list):
            option_items = []
        option_additionals = row.get("optionAdditionalAmounts")
        if not isinstance(option_additionals, list):
            option_additionals = []
        entries.append({
            "queueKey": text_value(row.get("queueKey") or f"{channel}:{gs}"),
            "account": account,
            "market": market,
            "channel": channel,
            "gs": gs,
            "sourceName": text_value(row.get("sourceName", "")),
            "title": text_value(row.get("title", "")),
            "searchTerms": text_value(row.get("searchTerms", "")),
            "mainImage": text_value(row.get("mainImage", "")),
            "mainImageSrc": text_value(row.get("mainImageSrc", "")),
            "additionalImageSrcs": [text_value(url) for url in additional_image_srcs if text_value(url)],
            "detailImageSrcs": [text_value(url) for url in detail_image_srcs if text_value(url)],
            "detailHtml": text_value(row.get("detailHtml", "")),
            "cafe24Url": text_value(row.get("cafe24Url", "")),
            "price": text_value(row.get("price") or row.get("salePrice") or ""),
            "supplyPrice": text_value(row.get("supplyPrice") or ""),
            "salePrice": text_value(row.get("salePrice") or row.get("price") or ""),
            "consumerPrice": text_value(row.get("consumerPrice") or ""),
            "optionSummary": text_value(row.get("optionSummary") or row.get("opt") or ""),
            "optionInput": text_value(row.get("optionInput") or ""),
            "optionAdditionalAmounts": option_additionals,
            "optionItems": option_items,
            "naverProvidedNotice": row.get("naverProvidedNotice") if isinstance(row.get("naverProvidedNotice"), dict) else {},
        })
    return entries


def parse_upload_price(value: object) -> int:
    text = text_value(value)
    if not text:
        return 0
    cleaned = re.sub(r"[^0-9.]", "", text)
    if not cleaned:
        return 0
    try:
        return int(float(cleaned))
    except Exception:
        return 0


def option_summary_to_input(option_summary: object) -> str:
    text = text_value(option_summary)
    if not text or "단일" in text:
        return ""
    body = text
    if "·" in body:
        body = body.split("·", 1)[1]
    body = re.sub(r"^\s*\d+\s*옵션\s*[:：-]?\s*", "", body).strip()
    parts = [
        re.sub(r"\s+", " ", part).strip(" ,/|")
        for part in re.split(r"\s*/\s*|\s*\|\s*|,\s*", body)
    ]
    parts = [part for part in parts if part and part not in {"옵션", "선택형"}]
    if not parts:
        return ""
    return "옵션{" + "|".join(f"{chr(65 + index)} {part}" for index, part in enumerate(parts[:26])) + "}"


def option_input_from_entry(entry: dict) -> str:
    existing = text_value(entry.get("optionInput"))
    if existing:
        return existing
    items = entry.get("optionItems") if isinstance(entry.get("optionItems"), list) else []
    labels = [
        text_value(item.get("option") if isinstance(item, dict) else item)
        for item in items
    ]
    labels = [label for label in labels if label and label != "단일"]
    if labels:
        return "옵션{" + "|".join(f"{chr(65 + index)} {label}" for index, label in enumerate(labels[:26])) + "}"
    return option_summary_to_input(entry.get("optionSummary"))


def option_additionals_from_entry(entry: dict) -> str:
    existing = entry.get("optionAdditionalAmounts")
    if isinstance(existing, list) and existing:
        return "|".join(str(int(parse_upload_price(value))) for value in existing)
    items = entry.get("optionItems") if isinstance(entry.get("optionItems"), list) else []
    if len(items) <= 1:
        return ""
    prices = [
        parse_upload_price(item.get("salePrice") or item.get("price"))
        for item in items
        if isinstance(item, dict)
    ]
    if len(prices) <= 1 or not prices[0]:
        return ""
    base = prices[0]
    return "|".join(str(price - base) for price in prices)


def upload_image_src_to_path(src: object) -> Path | None:
    raw = text_value(src)
    if not raw:
        return None
    parsed = urllib.parse.urlparse(raw)
    path_text = ""
    if parsed.scheme in {"http", "https"}:
        if parsed.path.startswith("/data/"):
            path_text = urllib.parse.unquote(parsed.path.lstrip("/"))
    elif raw.startswith("/data/"):
        path_text = urllib.parse.unquote(raw.lstrip("/"))
    elif raw.startswith("data/"):
        path_text = urllib.parse.unquote(raw)
    elif Path(raw).is_absolute():
        path = Path(raw)
        return path.resolve() if path.exists() else None

    if not path_text:
        return None
    path = (ROOT / path_text).resolve()
    return path if path.exists() and is_within(DATA_ROOT, path) else None


def public_image_url(src: object) -> str:
    raw = text_value(src)
    if raw.startswith("http://") or raw.startswith("https://"):
        parsed = urllib.parse.urlparse(raw)
        file_name = Path(urllib.parse.unquote(parsed.path)).name
        if file_name and not re.match(r"^[A-Za-z0-9._%+\-=()]+$", file_name):
            return ""
        return raw
    return ""


def direct_upload_image_ref(src: object) -> str:
    path = upload_image_src_to_path(src)
    if path is not None:
        return str(path)
    return public_image_url(src)


def seed_detail_images_for_gs(gs: object) -> list[str]:
    gs_value = text_value(gs).upper()
    if not gs_value:
        return []
    base_value = base_gs_code(gs_value)
    for seed_path in sorted(SEED_ROOT.glob("*.webseed.json"), key=lambda path: path.stat().st_mtime, reverse=True):
        seed = read_json(seed_path, {})
        products = seed.get("products")
        if not isinstance(products, list):
            continue
        for product in products:
            if not isinstance(product, dict):
                continue
            product_gs = text_value(product.get("gs")).upper()
            product_base = text_value(product.get("baseGs") or base_gs_code(product_gs)).upper()
            if gs_value not in {product_gs, product_base} and base_value not in {product_gs, product_base}:
                continue
            images = product.get("images")
            if not isinstance(images, dict):
                return []
            detail = images.get("detail")
            if isinstance(detail, list):
                return [text_value(url) for url in detail if text_value(url)]
            return extract_image_urls(detail)
    return []


def seed_detail_html_for_gs(gs: object) -> str:
    gs_value = text_value(gs).upper()
    if not gs_value:
        return ""
    base_value = base_gs_code(gs_value)
    for seed_path in sorted(SEED_ROOT.glob("*.webseed.json"), key=lambda path: path.stat().st_mtime, reverse=True):
        seed = read_json(seed_path, {})
        products = seed.get("products")
        if not isinstance(products, list):
            continue
        for product in products:
            if not isinstance(product, dict):
                continue
            product_gs = text_value(product.get("gs")).upper()
            product_base = text_value(product.get("baseGs") or base_gs_code(product_gs)).upper()
            if gs_value not in {product_gs, product_base} and base_value not in {product_gs, product_base}:
                continue
            return text_value(product.get("detailHtml"))
    return ""


def base_gs_code(gs: object) -> str:
    value = text_value(gs).upper()
    match = GS_CODE_RE.search(value)
    return match.group(1).upper() if match else value


def image_selection_for_entry(entry: dict) -> tuple[int | None, list[int]]:
    main_path = upload_image_src_to_path(entry.get("mainImageSrc"))
    add_paths = [upload_image_src_to_path(url) for url in entry.get("additionalImageSrcs", [])]
    paths = [path for path in [main_path, *add_paths] if path is not None]
    if not paths:
        return None, []
    folder = paths[0].parent
    files = [
        path.resolve()
        for path in sorted(folder.iterdir(), key=image_file_sort_key)
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
    ]
    if not files:
        return None, []
    index_by_path = {str(path): index for index, path in enumerate(files)}
    main_index = index_by_path.get(str(main_path.resolve())) if main_path else None
    additional: list[int] = []
    for path in add_paths:
        if path is None:
            continue
        index = index_by_path.get(str(path.resolve()))
        if index is not None and index != main_index and index not in additional:
            additional.append(index)
    return main_index, additional


def update_export_image_selections(entries: list[dict]) -> None:
    if not entries:
        return
    target = EXPORT_ROOT / "image_selections.json"
    payload = read_json(target, {})
    if not isinstance(payload, dict):
        payload = {}
    nested = payload.get("image_selections")
    if not isinstance(nested, dict):
        nested = {}
    changed = False
    for entry in entries:
        gs_key = base_gs_code(entry.get("gs"))
        if not gs_key:
            continue
        main_index, additional = image_selection_for_entry(entry)
        if main_index is None:
            continue
        selection = {"main": main_index, "additional": additional}
        payload[gs_key[:9]] = selection
        nested[gs_key[:9]] = selection
        changed = True
    if changed:
        payload["image_selections"] = nested
        write_json(target, payload)


def infer_direct_upload_categories(entry: dict) -> dict[str, str]:
    text = " ".join(
        text_value(entry.get(key))
        for key in ("title", "sourceName", "searchTerms", "optionSummary", "gs")
    )
    compact = re.sub(r"\s+", "", text)
    out: dict[str, str] = {}

    def apply(**values: str) -> None:
        for key, value in values.items():
            if value and not out.get(key):
                out[key] = value

    if re.search(r"깔창|인솔|신발밑창|밑창보강", compact, re.IGNORECASE):
        apply(
            naver="50000667",
            coupang="64623",
            lotte_standard="BC43071000",
            lotte_display="FC18101001",
            lotte_item="38",
        )
    if re.search(r"카라비너|릴고리|릴홀더|키홀더|키링|연결고리|등산고리|고리", compact, re.IGNORECASE):
        apply(
            naver="50002646",
            coupang="81718",
            lotte_standard="BC20040800",
            lotte_display="EC10400324",
            lotte_item="38",
        )
    if re.search(r"에어컨.*커버|실외기.*커버|커버.*실외기", compact, re.IGNORECASE):
        apply(
            naver="50003518",
            coupang="78137",
            lotte_standard="BC63120300",
            lotte_display="FC11160703",
            lotte_item="38",
        )
    if re.search(r"가구발커버|의자발커버|가구커버|소파커버|커버류", compact, re.IGNORECASE):
        apply(
            naver="50003521",
            coupang="78133",
            lotte_standard="BC63120300",
            lotte_display="FC11160703",
            lotte_item="38",
        )
    if re.search(r"작업장갑|안전장갑|코팅장갑|나일론.*장갑|오픈핑거.*장갑|장갑", compact, re.IGNORECASE):
        apply(
            naver="50003450",
            coupang="64387",
            lotte_standard="BC10040800",
            lotte_display="FC19041003",
            lotte_item="04",
        )
    if re.search(r"나사|스크류|볼트|너트|브라켓|고정핀|철물|부속|부품", compact, re.IGNORECASE):
        apply(
            naver="50003466",
            coupang="64310",
            lotte_standard="BC10080200",
            lotte_display="FC19040401",
            lotte_item="38",
        )
    return out


def upload_workbook_headers() -> list[str]:
    return [
        "상품코드", "자체 상품코드", "판매자내부상품번호", "상품명", "공급사 상품명",
        "홈런_공통마켓상품명", "홈런_네이버상품명", "네이버상품명", "홈런_네이버태그", "네이버태그",
        "홈런_쿠팡상품명", "쿠팡상품명", "홈런_쿠팡검색태그", "쿠팡검색태그",
        "홈런_롯데ON상품명", "롯데ON상품명", "홈런_롯데ON검색키워드", "롯데ON검색키워드",
        "네이버카테고리코드", "쿠팡카테고리코드", "롯데ON표준카테고리코드", "롯데ON전시카테고리코드", "롯데ON상품품목코드",
        "11번가카테고리코드", "11번가카테고리코드/경로",
        "ESM카테고리코드", "ESM카테고리코드/경로", "옥션카테고리코드", "G마켓카테고리코드",
        "홈런_공통마켓검색키워드", "공통마켓검색키워드", "검색어설정", "검색키워드",
        "판매가", "상품가", "소비자가", "옵션입력", "옵션추가금",
        "이미지등록(목록)", "이미지등록(추가)", "이미지등록(상세)",
        "상품 상세설명", "상세설명", "브랜드",
    ]


def build_upload_workbook_row(entry: dict) -> dict[str, object]:
    title = clean_product_title(entry.get("title")) or clean_product_title(entry.get("sourceName")) or text_value(entry.get("gs"))
    search_terms = text_value(entry.get("searchTerms"))
    price = parse_upload_price(entry.get("salePrice") or entry.get("price")) or 1000
    consumer_price = parse_upload_price(entry.get("consumerPrice")) or round_100(price * 1.2) or price
    upload_images = [
        url for url in [
            direct_upload_image_ref(entry.get("mainImageSrc")),
            *[direct_upload_image_ref(url) for url in entry.get("additionalImageSrcs", [])],
        ]
        if url
    ]
    public_images = [
        url for url in [
            public_image_url(entry.get("mainImageSrc")),
            *[public_image_url(url) for url in entry.get("additionalImageSrcs", [])],
        ]
        if url
    ]
    detail_sources = entry.get("detailImageSrcs") if isinstance(entry.get("detailImageSrcs"), list) else []
    if not detail_sources:
        detail_sources = seed_detail_images_for_gs(entry.get("gs"))
    public_detail_images = [
        public_image_url(url)
        for url in detail_sources
        if public_image_url(url)
    ]
    image_list = "|".join(upload_images[:9])
    detail_html = text_value(entry.get("detailHtml")) or seed_detail_html_for_gs(entry.get("gs"))
    if detail_html and "<img" not in detail_html.lower() and public_detail_images:
        detail_html = ""
    if not detail_html and public_detail_images:
        detail_html = "<center>" + "".join(f'<img src="{url}">' for url in public_detail_images[:80]) + "</center>"
    elif not detail_html and public_images:
        detail_html = "<center>" + "".join(f'<img src="{url}">' for url in public_images[:12]) + "</center>"
    if not detail_html:
        detail_html = "상세페이지 참조"

    categories = infer_direct_upload_categories(entry)
    return {
        "상품코드": entry["gs"],
        "자체 상품코드": entry["gs"],
        "판매자내부상품번호": entry["gs"],
        "상품명": title,
        "공급사 상품명": title,
        "홈런_공통마켓상품명": title,
        "홈런_네이버상품명": title if entry["market"] == "네이버" else "",
        "네이버상품명": title if entry["market"] == "네이버" else "",
        "홈런_네이버태그": search_terms if entry["market"] == "네이버" else "",
        "네이버태그": search_terms if entry["market"] == "네이버" else "",
        "홈런_쿠팡상품명": title if entry["market"] == "쿠팡" else "",
        "쿠팡상품명": title if entry["market"] == "쿠팡" else "",
        "홈런_쿠팡검색태그": search_terms if entry["market"] == "쿠팡" else "",
        "쿠팡검색태그": search_terms if entry["market"] == "쿠팡" else "",
        "홈런_롯데ON상품명": title if entry["market"] == "롯데ON" else "",
        "롯데ON상품명": title if entry["market"] == "롯데ON" else "",
        "홈런_롯데ON검색키워드": search_terms if entry["market"] == "롯데ON" else "",
        "롯데ON검색키워드": search_terms if entry["market"] == "롯데ON" else "",
        "네이버카테고리코드": categories.get("naver", ""),
        "쿠팡카테고리코드": categories.get("coupang", ""),
        "롯데ON표준카테고리코드": categories.get("lotte_standard", ""),
        "롯데ON전시카테고리코드": categories.get("lotte_display", ""),
        "롯데ON상품품목코드": categories.get("lotte_item", ""),
        "11번가카테고리코드": categories.get("elevenst", ""),
        "11번가카테고리코드/경로": categories.get("elevenst", ""),
        "ESM카테고리코드": categories.get("esm", ""),
        "ESM카테고리코드/경로": categories.get("esm", ""),
        "옥션카테고리코드": categories.get("auction", ""),
        "G마켓카테고리코드": categories.get("gmarket", ""),
        "홈런_공통마켓검색키워드": search_terms,
        "공통마켓검색키워드": search_terms,
        "검색어설정": search_terms,
        "검색키워드": search_terms,
        "판매가": price,
        "상품가": price,
        "소비자가": consumer_price,
        "옵션입력": option_input_from_entry(entry),
        "옵션추가금": option_additionals_from_entry(entry),
        "이미지등록(목록)": image_list,
        "이미지등록(추가)": "|".join(upload_images[1:9]),
        "이미지등록(상세)": image_list,
        "상품 상세설명": detail_html,
        "상세설명": detail_html,
        "브랜드": "샤플라이",
    }


def write_direct_upload_workbook(job_id: str, entry: dict) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl 로드 실패: {exc}") from exc

    file_name = safe_name(f"api_upload_{job_id}_{entry['channel']}_{entry['gs']}.xlsx")
    target = EXPORT_ROOT / file_name
    headers = upload_workbook_headers()
    row = build_upload_workbook_row(entry)

    workbook = Workbook()
    first = True
    for sheet_name in ("분리추출후", "A마켓", "B마켓"):
        sheet = workbook.active if first else workbook.create_sheet(sheet_name)
        first = False
        sheet.title = sheet_name
        sheet.append(headers)
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(target)
    return target


def market_cli_flag(market: str) -> str:
    return {
        "네이버": "--naver",
        "쿠팡": "--coupang",
        "롯데ON": "--lotteon",
    }.get(market, "")


def resolve_market_key_path(settings: dict, account: str, market: str) -> Path:
    item = settings.get(market_key_id(account, market))
    if not isinstance(item, dict):
        raise FileNotFoundError(f"{account}:{market} 키 파일이 없습니다.")
    path = Path(item.get("path", "")).resolve()
    if not path.is_file() or not is_within(MARKET_KEY_ROOT, path):
        raise FileNotFoundError(f"{account}:{market} 키 파일을 찾지 못했습니다.")
    return path


def backup_or_remove_key_file(target: Path, temp_dir: Path) -> tuple[str, Path | None]:
    if target.exists():
        backup = temp_dir / target.name
        shutil.copy2(target, backup)
        return "backup", backup
    return "missing", None


@contextmanager
def market_key_overlay(account: str, market: str, settings: dict):
    DESKTOP_KEY_ROOT.mkdir(parents=True, exist_ok=True)
    temp_dir = Path(tempfile.mkdtemp(prefix="webocr_key_overlay_", dir=str(EXPORT_ROOT)))
    restored: list[tuple[Path, str, Path | None]] = []

    def overlay_file(source: Path, target_name: str, hide_only: bool = False) -> None:
        target = (DESKTOP_KEY_ROOT / target_name).resolve()
        if source.resolve() == target:
            return
        state, backup = backup_or_remove_key_file(target, temp_dir)
        restored.append((target, state, backup))
        if hide_only:
            if target.exists():
                target.unlink()
            return
        shutil.copy2(source, target)

    with MARKET_KEY_OVERLAY_LOCK:
        try:
            source = resolve_market_key_path(settings, account, market)
            if market == "네이버":
                overlay_file(source, "naver_client_key.txt")
            elif market == "쿠팡":
                overlay_file(source, "coupang_wing_api.txt")
            elif market == "롯데ON":
                if source.suffix.lower() == ".json":
                    overlay_file(source, "lotteon_upload_id.json")
                else:
                    overlay_file(source, "lotteon_api.txt")
                    overlay_file(source, "lotteon_upload_id.json", hide_only=True)
            elif market == "Cafe24":
                pass
            else:
                raise ValueError(f"{market}은 API 업로드 대상이 아닙니다.")

            cafe24_item = settings.get(market_key_id(account, "Cafe24"))
            if isinstance(cafe24_item, dict):
                cafe24_path = Path(cafe24_item.get("path", "")).resolve()
                if cafe24_path.is_file() and is_within(MARKET_KEY_ROOT, cafe24_path):
                    overlay_file(cafe24_path, "cafe24_token_rkghrud1.json")
                    overlay_file(cafe24_path, "cafe24_token.json")
                    if account == "B":
                        overlay_file(cafe24_path, "cafe24_token_jb.json")
            yield
        finally:
            for target, state, backup in reversed(restored):
                try:
                    if state == "backup" and backup and backup.exists():
                        shutil.copy2(backup, target)
                    elif state == "missing" and target.exists():
                        target.unlink()
                except Exception:
                    pass
            shutil.rmtree(temp_dir, ignore_errors=True)


def parse_direct_upload_result(entry: dict, exit_code: int, lines: list[str]) -> dict:
    result = {**entry, "status": "failed", "updatedAt": now_text(), "error": "", "productId": ""}
    tail_text = "\n".join(lines[-80:])
    if exit_code != 0:
        result["error"] = scrub_secret(tail_text or f"dotnet exit code {exit_code}", 400)
        return result

    market = entry["market"]
    if market == "네이버":
        pattern = r"\[네이버\]\s+row=\d+\s+status=(\S+)\s+id=(.*?)\s+error=(.*)"
        id_name = "productId"
    elif market == "쿠팡":
        pattern = r"\[쿠팡\]\s+row=\d+\s+status=(\S+)\s+id=(.*?)\s+category=.*?\s+error=(.*)"
        id_name = "sellerProductId"
    else:
        pattern = r"\[롯데ON\]\s+row=\d+\s+status=(\S+)\s+spdNo=(.*?)\s+error=(.*)"
        id_name = "spdNo"

    matches = re.findall(pattern, tail_text)
    if not matches:
        result["error"] = scrub_secret(tail_text or "업로드 결과를 파싱하지 못했습니다.", 400)
        return result

    status, product_id, error = matches[-1]
    status_upper = status.upper()
    result[id_name] = product_id.strip()
    result["productId"] = product_id.strip()
    result["rawStatus"] = status_upper
    if status_upper in {"OK", "SUCCESS", "DRY_RUN_OK", "DRY_RUN"}:
        result["status"] = "uploaded"
    elif status_upper in {"SKIP_DUP"}:
        result["status"] = "skipped"
        result["error"] = error.strip() or "기존 성공 이력 있음"
    else:
        result["status"] = "failed"
        result["error"] = error.strip() or status_upper
    return result


def execute_direct_market_upload(job_id: str, entry: dict, workbook_path: Path, log, results: list[dict]) -> tuple[dict, list[str]]:
    flag = market_cli_flag(entry["market"])
    if not flag:
        raise ValueError(f"{entry['market']}은 API 업로드 대상이 아닙니다.")
    cmd = [
        "dotnet", "run",
        "--project", str(DOTNET_UPLOAD_PROJECT),
        "--",
        "--direct-market-upload",
        "--file", str(workbook_path),
        "--gs", entry["gs"],
        flag,
    ]
    if entry["market"] == "롯데ON":
        cmd.append("--force")
    if entry.get("dryRun"):
        cmd.append("--dry-run")

    log.write(" ".join(f'"{part}"' if " " in part else part for part in cmd) + "\n")
    process = subprocess.Popen(
        cmd,
        cwd=str(PROJECT_ROOT),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
    )
    register_active_process(job_id, process)
    lines: list[str] = []
    output_queue: queue.Queue[str | None] = queue.Queue()

    def read_stdout() -> None:
        try:
            assert process.stdout is not None
            for raw_line in process.stdout:
                output_queue.put(raw_line)
        finally:
            output_queue.put(None)

    reader = threading.Thread(target=read_stdout, daemon=True)
    reader.start()
    started = time.monotonic()
    exit_code = -1
    try:
        stream_closed = False
        while True:
            if job_id in CANCELLED_JOB_IDS:
                stop_process_tree(process)
                exit_code = -1
                break
            if time.monotonic() - started > MARKET_UPLOAD_TIMEOUT_SECONDS:
                lines.append(f"TIMEOUT: {entry['channel']} {entry['gs']} API 업로드가 {MARKET_UPLOAD_TIMEOUT_SECONDS}초를 초과했습니다.")
                stop_process_tree(process)
                exit_code = -1
                break
            try:
                item = output_queue.get(timeout=0.5)
            except queue.Empty:
                if process.poll() is not None and stream_closed:
                    exit_code = process.returncode
                    break
                continue
            if item is None:
                stream_closed = True
                if process.poll() is not None:
                    exit_code = process.returncode
                    break
                continue
            clean = item.rstrip("\n")
            if clean:
                lines.append(clean)
                log.write(clean + "\n")
                log.flush()
                current = read_json(JOBS_ROOT / f"{job_id}.json", {"jobId": job_id})
                current.update({
                    "status": "running",
                    "updatedAt": now_text(),
                    "currentStage": f"{entry['channel']} API 실행",
                    "tail": lines[-20:],
                    "results": results,
                })
                write_json(JOBS_ROOT / f"{job_id}.json", current)
        if exit_code == -1 and process.poll() is not None and not any(line.startswith("TIMEOUT:") for line in lines):
            exit_code = process.returncode
    finally:
        unregister_active_process(job_id, process)
    return parse_direct_upload_result(entry, exit_code, lines), lines


def run_market_upload_job(job_id: str, payload: dict) -> None:
    job_path = JOBS_ROOT / f"{job_id}.json"
    log_path = JOBS_ROOT / f"{job_id}.log"
    entries = normalize_upload_entries(payload)
    dry_run = bool(payload.get("dryRun"))
    if dry_run:
        for entry in entries:
            entry["dryRun"] = True
    job = read_json(job_path, {"jobId": job_id})
    job.update({
        "status": "running",
        "startedAt": now_text(),
        "logPath": str(log_path),
        "progressPercent": 5,
        "currentStage": "업로드 요청 검증",
        "total": len(entries),
    })
    write_json(job_path, job)

    settings = read_market_key_settings().get("items", {})
    results: list[dict] = []
    api_markets = {"네이버", "쿠팡", "롯데ON"}
    try:
        update_export_image_selections(entries)
        with log_path.open("w", encoding="utf-8", errors="replace") as log:
            log.write(f"[{now_text()}] START marketUpload\n")
            log.write(f"entries: {len(entries)}\n")
            if not entries:
                raise ValueError("upload entries empty")
            for index, entry in enumerate(entries, start=1):
                if job_id in CANCELLED_JOB_IDS:
                    raise RuntimeError("업로드 작업이 중지되었습니다.")
                market = entry["market"]
                key_id = market_key_id(entry["account"], market)
                result = {**entry, "status": "failed", "updatedAt": now_text()}
                if market not in api_markets:
                    result["status"] = "skipped"
                    result["error"] = "API 업로드 대상이 아닙니다. 11번가/ESM은 엑셀 export를 사용합니다."
                elif key_id not in settings:
                    result["error"] = f"{entry['channel']} 키 파일이 없습니다."
                else:
                    result["keyFile"] = settings[key_id].get("fileName", "")
                    workbook_path = write_direct_upload_workbook(job_id, entry)
                    result["workbookPath"] = str(workbook_path)
                    log.write(f"[{index}/{len(entries)}] {entry['channel']} {entry['gs']} 실제 API 업로드 시작\n")
                    log.flush()
                    with market_key_overlay(entry["account"], market, settings):
                        result, _lines = execute_direct_market_upload(job_id, entry, workbook_path, log, results)
                    result["keyFile"] = settings[key_id].get("fileName", "")
                    result["workbookPath"] = str(workbook_path)
                results.append(result)
                log.write(f"[{index}/{len(entries)}] {entry['channel']} {entry['gs']} -> {result['status']} {result.get('error', '')}\n")
                log.flush()
                job = read_json(job_path, job)
                job.update({
                    "status": "running",
                    "updatedAt": now_text(),
                    "progressPercent": min(95, 10 + int(index / max(len(entries), 1) * 80)),
                    "currentStage": f"API 업로드 {index}/{len(entries)}",
                    "results": results,
                    "tail": [f"{item['channel']} {item['gs']} {item['status']}" for item in results[-20:]],
                })
                write_json(job_path, job)
        job.update({
            "status": "completed",
            "finishedAt": now_text(),
            "progressPercent": 100,
            "currentStage": "API 업로드 완료",
            "result": {
                "total": len(results),
                "success": sum(1 for item in results if item.get("status") == "uploaded"),
                "failed": sum(1 for item in results if item.get("status") == "failed"),
                "skipped": sum(1 for item in results if item.get("status") == "skipped"),
                "results": results,
            },
        })
        write_json(job_path, job)
    except Exception as exc:
        with log_path.open("a", encoding="utf-8", errors="replace") as log:
            log.write(f"\n[{now_text()}] ERROR {exc}\n")
        job.update({"status": "failed", "finishedAt": now_text(), "error": str(exc), "currentStage": "업로드 실패"})
        write_json(job_path, job)


def find_active_market_upload_job() -> str:
    for path in sorted(JOBS_ROOT.glob("*.json"), key=lambda item: item.stat().st_mtime, reverse=True):
        try:
            if time.time() - path.stat().st_mtime > 7200:
                continue
            job = read_json(path, {})
        except Exception:
            continue
        if job.get("action") == "marketUpload" and job.get("status") in {"queued", "running"}:
            return text_value(job.get("jobId") or path.stem)
    return ""


def export_file_url(path: Path) -> str:
    try:
        rel = path.resolve().relative_to(EXPORT_ROOT.resolve())
        rel_text = str(rel).replace("\\", "/")
        return f"/data/exports/{urllib.parse.quote(rel_text, safe='/')}"
    except Exception:
        return f"/data/exports/{urllib.parse.quote(path.name)}"


def write_official_market_export_source(entries: list[dict], market: str, stamp: str) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl 로드 실패: {exc}") from exc

    source_dir = EXPORT_ROOT / safe_name(f"{market}_official_source_{stamp}", "official_source")
    source_dir.mkdir(parents=True, exist_ok=True)
    listing_root = source_dir / "listing_images"
    listing_root.mkdir(parents=True, exist_ok=True)
    source_path = source_dir / "market_export_source.xlsx"
    headers = upload_workbook_headers()

    selection_payload: dict[str, object] = {}
    nested: dict[str, object] = {}

    for entry in entries:
        gs_base = base_gs_code(entry.get("gs"))
        if not gs_base:
            continue
        gs_folder = listing_root / gs_base
        gs_folder.mkdir(parents=True, exist_ok=True)
        image_sources = [
            entry.get("mainImageSrc"),
            *(
                entry.get("additionalImageSrcs")
                if isinstance(entry.get("additionalImageSrcs"), list)
                else []
            ),
        ]
        copied_count = 0
        for index, src in enumerate(image_sources):
            source_path_value = upload_image_src_to_path(src)
            if source_path_value is None or not source_path_value.exists():
                continue
            suffix = source_path_value.suffix.lower()
            if suffix not in IMAGE_EXTENSIONS:
                suffix = ".jpg"
            name = f"{index:03d}_{'main' if index == 0 else 'add'}{suffix}"
            shutil.copy2(source_path_value, gs_folder / name)
            copied_count += 1
        if copied_count > 0:
            selection = {"main": 0, "additional": list(range(1, copied_count))}
            selection_payload[gs_base[:9]] = selection
            nested[gs_base[:9]] = selection

    if nested:
        selection_payload["image_selections"] = nested
        write_json(source_dir / "image_selections.json", selection_payload)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "분리추출후"
    sheet.append(headers)
    for entry in entries:
        row = build_upload_workbook_row(entry)
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(source_path)
    return source_path


def parse_market_excel_files(lines: list[str]) -> list[dict]:
    result: list[dict] = []
    known_markets = ["11번가 이미지ZIP", "11번가", "ESM"]
    for line in lines:
        if not line.startswith("MARKET_EXCEL_FILE "):
            continue
        tail = line.replace("MARKET_EXCEL_FILE ", "", 1).strip()
        for market in known_markets:
            prefix = f"{market} "
            if tail.startswith(prefix):
                path = Path(tail[len(prefix):].strip())
                result.append({
                    "market": market,
                    "fileName": path.name,
                    "path": str(path),
                    "url": export_file_url(path),
                })
                break
    return result


def write_official_market_export(payload: dict, entries: list[dict], market: str, stamp: str) -> dict:
    source_path = write_official_market_export_source(entries, market, stamp)
    selected_gs = ",".join(sorted({text_value(entry.get("gs")) for entry in entries if text_value(entry.get("gs"))}))
    if not selected_gs:
        raise ValueError("export GS list empty")
    account = text_value(entries[0].get("account")).upper() or "A"
    settings = read_market_key_settings().get("items", {})
    cmd = [
        "dotnet", "run",
        "--project", str(DOTNET_UPLOAD_PROJECT),
        "--",
        "--export-market-excel",
        "--source", str(source_path),
        "--gs", selected_gs,
        "--project-root", str(PROJECT_ROOT),
    ]
    cafe24_key = market_key_id(account, "Cafe24")
    overlay_context = market_key_overlay(account, "Cafe24", settings) if cafe24_key in settings else nullcontext()
    with overlay_context:
        completed = subprocess.run(
            cmd,
            cwd=str(PROJECT_ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=MARKET_UPLOAD_TIMEOUT_SECONDS,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )

    lines = [
        line.strip()
        for line in (completed.stdout + "\n" + completed.stderr).splitlines()
        if line.strip()
    ]
    if completed.returncode != 0:
        raise RuntimeError(scrub_secret("\n".join(lines[-40:]) or f"dotnet exit code {completed.returncode}", 1200))

    files = parse_market_excel_files(lines)
    primary = next((file for file in files if file.get("market") == market), None)
    if not primary:
        raise RuntimeError("공식 마켓 엑셀 결과 파일을 찾지 못했습니다.")
    related = [
        file for file in files
        if file.get("market") == market or (market == "11번가" and file.get("market") == "11번가 이미지ZIP")
    ]
    return {
        **primary,
        "count": len(entries),
        "format": Path(primary["path"]).suffix.lstrip(".") or "xlsx",
        "sourcePath": str(source_path),
        "files": related,
        "officialTemplate": True,
        "stdoutTail": lines[-20:],
    }


def write_market_export(payload: dict) -> dict:
    entries = normalize_upload_entries(payload)
    market = safe_name(text_value(payload.get("market") or "market"), "market")
    if not entries:
        raise ValueError("export entries empty")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if market in {"11번가", "ESM"}:
        try:
            return write_official_market_export(payload, entries, market, stamp)
        except Exception as exc:
            fallback_reason = str(exc)
    else:
        fallback_reason = ""
    headers = [
        "계정", "마켓", "GS코드", "원본상품명", "업로드상품명", "검색어설정",
        "판매가", "소비자가", "옵션입력", "옵션추가금",
        "대표이미지", "이미지등록(목록)", "이미지등록(추가)", "이미지등록(상세)",
        "상품 상세설명", "Cafe24 URL", "상태",
    ]
    try:
        from openpyxl import Workbook

        file_name = safe_name(f"{market}_upload_queue_{stamp}.xlsx")
        target = EXPORT_ROOT / file_name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = market[:31]
        sheet.append(headers)
        for entry in entries:
            price = parse_upload_price(entry.get("salePrice") or entry.get("price"))
            consumer_price = parse_upload_price(entry.get("consumerPrice")) or round_100(price * 1.2)
            main_image = direct_upload_image_ref(entry.get("mainImageSrc"))
            add_images = [direct_upload_image_ref(url) for url in entry.get("additionalImageSrcs", [])]
            add_images = [url for url in add_images if url]
            detail_html = text_value(entry.get("detailHtml")) or seed_detail_html_for_gs(entry.get("gs"))
            if not detail_html:
                detail_urls = [public_image_url(url) for url in entry.get("detailImageSrcs", []) if public_image_url(url)]
                if detail_urls:
                    detail_html = "<center>" + "".join(f'<img src="{url}">' for url in detail_urls[:80]) + "</center>"
            sheet.append([
                entry["account"],
                entry["market"],
                entry["gs"],
                entry["sourceName"],
                entry["title"],
                entry["searchTerms"],
                price,
                consumer_price,
                option_input_from_entry(entry),
                option_additionals_from_entry(entry),
                entry["mainImage"],
                main_image,
                main_image,
                "\n".join(add_images),
                "|".join([main_image, *add_images]) if main_image else "",
                detail_html,
                entry["cafe24Url"],
                "대기",
            ])
        workbook.save(target)
        file_format = "xlsx"
    except Exception:
        file_name = safe_name(f"{market}_upload_queue_{stamp}.csv")
        target = EXPORT_ROOT / file_name
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for entry in entries:
                price = parse_upload_price(entry.get("salePrice") or entry.get("price"))
                consumer_price = parse_upload_price(entry.get("consumerPrice")) or round_100(price * 1.2)
                main_image = direct_upload_image_ref(entry.get("mainImageSrc"))
                add_images = [direct_upload_image_ref(url) for url in entry.get("additionalImageSrcs", [])]
                add_images = [url for url in add_images if url]
                detail_html = text_value(entry.get("detailHtml")) or seed_detail_html_for_gs(entry.get("gs"))
                writer.writerow([
                    entry["account"],
                    entry["market"],
                    entry["gs"],
                    entry["sourceName"],
                    entry["title"],
                    entry["searchTerms"],
                    price,
                    consumer_price,
                    option_input_from_entry(entry),
                    option_additionals_from_entry(entry),
                    entry["mainImage"],
                    main_image,
                    main_image,
                    "\n".join(add_images),
                    "|".join([main_image, *add_images]) if main_image else "",
                    detail_html,
                    entry["cafe24Url"],
                    "대기",
                ])
        file_format = "csv"
    return {
        "fileName": file_name,
        "path": str(target),
        "url": export_file_url(target),
        "count": len(entries),
        "format": file_format,
        "officialTemplate": False,
        "fallbackReason": fallback_reason,
    }


def remove_runtime_path(raw_path: object) -> dict:
    path_text = text_value(raw_path)
    if not path_text:
        return {"path": "", "status": "skipped", "reason": "empty"}
    path = Path(path_text)
    if not path.is_absolute():
        path = (ROOT / path).resolve()
    else:
        path = path.resolve()
    allowed_roots = [UPLOAD_ROOT, JOBS_ROOT, SEED_ROOT, EXPORT_ROOT]
    if not any(is_within(root, path) for root in allowed_roots):
        return {"path": str(path), "status": "skipped", "reason": "outside runtime data"}
    if not path.exists():
        return {"path": str(path), "status": "missing"}
    if path.is_dir():
        shutil.rmtree(path)
    else:
        path.unlink()
    return {"path": str(path), "status": "deleted"}


def cleanup_workspace_artifacts(payload: dict) -> dict:
    paths = payload.get("paths") if isinstance(payload.get("paths"), list) else []
    job_ids = payload.get("jobIds") if isinstance(payload.get("jobIds"), list) else []
    removed: list[dict] = []
    stopped: list[str] = []
    seen_paths: set[str] = set()

    def remember_path(path_value: object) -> None:
        clean = text_value(path_value)
        if clean and clean not in seen_paths:
            seen_paths.add(clean)
            removed.append(remove_runtime_path(clean))

    for raw_job_id in job_ids:
        job_id = safe_name(text_value(raw_job_id), "")
        if not job_id:
            continue
        if stop_active_job(job_id):
            stopped.append(job_id)
        job_path = JOBS_ROOT / f"{job_id}.json"
        job_payload = read_json(job_path, {})
        remember_path(job_path)
        remember_path(JOBS_ROOT / f"{job_id}.log")
        remember_path(JOBS_ROOT / f"{job_id}_keyword")
        result = job_payload.get("result") if isinstance(job_payload.get("result"), dict) else {}
        remember_path(result.get("keywordResultPath", ""))

    for path_value in paths:
        remember_path(path_value)

    return {
        "stoppedJobs": stopped,
        "removed": removed,
        "deleted": sum(1 for item in removed if item.get("status") == "deleted"),
        "skipped": sum(1 for item in removed if item.get("status") == "skipped"),
        "missing": sum(1 for item in removed if item.get("status") == "missing"),
    }


class WebOcrHandler(SimpleHTTPRequestHandler):
    server_version = "WEBOCRV2Local/0.1"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def log_message(self, format: str, *args) -> None:
        sys.stderr.write("%s - - [%s] %s\n" % (self.address_string(), self.log_date_time_string(), format % args))

    def send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def read_body(self) -> bytes:
        length = int(self.headers.get("Content-Length", "0") or "0")
        return self.rfile.read(length)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path == "/api/health":
            self.send_json({"ok": True, "time": now_text()})
            return
        if path == "/api/seeds":
            self.send_json({"ok": True, "seeds": list_seed_summaries()})
            return
        if path == "/api/market-keys":
            self.send_json({"ok": True, "items": market_key_summaries()})
            return
        if path == "/api/seed":
            params = parse_qs(parsed.query)
            seed_path = resolve_seed_path((params.get("path") or params.get("name") or [""])[0])
            if not seed_path.exists():
                self.send_json({"ok": False, "error": "seed not found"}, 404)
                return
            self.send_json({"ok": True, "seed": hydrate_seed_payload(read_json(seed_path, {})), "summary": seed_summary(seed_path)})
            return
        if path.startswith("/api/jobs/"):
            job_id = safe_name(path.rsplit("/", 1)[-1], "")
            job = read_json(JOBS_ROOT / f"{job_id}.json", {})
            if not job:
                self.send_json({"ok": False, "error": "job not found"}, 404)
                return
            self.send_json({"ok": True, **job})
            return
        return super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        try:
            if path == "/api/import-source":
                self.handle_import_file(UPLOAD_ROOT, "file")
                return
            if path == "/api/import-logo":
                self.handle_import_file(LOGO_ROOT, "file")
                return
            if path == "/api/market-key":
                self.handle_market_key_upload()
                return
            if path == "/api/market-key-test":
                self.handle_market_key_test()
                return
            if path == "/api/source-to-seed":
                self.handle_source_to_seed()
                return
            if path == "/api/keyword-generate":
                self.handle_keyword_generate()
                return
            if path == "/api/job-stop":
                self.handle_job_stop()
                return
            if path == "/api/market-upload":
                self.handle_market_upload()
                return
            if path == "/api/excel-export":
                self.handle_excel_export()
                return
            if path == "/api/workspace-reset":
                self.handle_workspace_reset()
                return
            if path == "/api/seed-action":
                self.handle_seed_action()
                return
            self.send_json({"ok": False, "error": "unknown endpoint"}, 404)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, 500)

    def handle_import_file(self, target_root: Path, field_name: str) -> None:
        body = self.read_body()
        parts = parse_multipart(body, self.headers.get("Content-Type", ""))
        file_part = parts.get(field_name)
        if not file_part or not file_part.get("filename"):
            self.send_json({"ok": False, "error": "file field missing"}, 400)
            return
        original = safe_name(file_part["filename"])
        upload_id = uuid.uuid4().hex[:12]
        target_path = target_root / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{upload_id}_{original}"
        target_path.write_bytes(file_part["content"])
        payload = {
            "ok": True,
            "uploadId": upload_id,
            "originalName": original,
            "path": str(target_path),
            "size": target_path.stat().st_size,
        }
        if target_root == UPLOAD_ROOT:
            payload["parsed"] = parse_source_preview(target_path)
        self.send_json(payload)

    def handle_market_key_upload(self) -> None:
        body = self.read_body()
        parts = parse_multipart(body, self.headers.get("Content-Type", ""))
        file_part = parts.get("file")
        if not file_part or not file_part.get("filename"):
            self.send_json({"ok": False, "error": "file field missing"}, 400)
            return
        account = text_value((parts.get("account") or {}).get("content", b"").decode("utf-8", errors="replace")).upper() or "A"
        market = text_value((parts.get("market") or {}).get("content", b"").decode("utf-8", errors="replace"))
        mode = text_value((parts.get("mode") or {}).get("content", b"").decode("utf-8", errors="replace")) or "key"
        if not market:
            self.send_json({"ok": False, "error": "market field missing"}, 400)
            return
        key = market_key_id(account, market)
        target_dir = MARKET_KEY_ROOT / safe_name(account) / safe_name(market)
        target_dir.mkdir(parents=True, exist_ok=True)
        original = safe_name(file_part["filename"])
        target_path = target_dir / original
        target_path.write_bytes(file_part["content"])

        settings = read_market_key_settings()
        settings["items"][key] = {
            "account": account,
            "market": market,
            "mode": mode,
            "fileName": original,
            "path": str(target_path),
            "updatedAt": now_text(),
        }
        write_market_key_settings(settings)
        self.send_json({
            "ok": True,
            "item": {
                "key": key,
                "account": account,
                "market": market,
                "mode": mode,
                "fileName": original,
                "size": target_path.stat().st_size,
                "updatedAt": settings["items"][key]["updatedAt"],
                "exists": True,
            },
            "items": market_key_summaries(),
        })

    def handle_market_key_test(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        settings = read_market_key_settings()
        items = settings.get("items", {})
        target_key = market_key_id(payload.get("account", ""), payload.get("market", "")) if payload.get("account") and payload.get("market") else ""
        selected = [item for key, item in items.items() if not target_key or key == target_key]
        results = [test_market_key(item) for item in selected]
        self.send_json({"ok": True, "results": results})

    def handle_source_to_seed(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        job_id = uuid.uuid4().hex[:12]
        job_path = JOBS_ROOT / f"{job_id}.json"
        job = {
            "ok": True,
            "jobId": job_id,
            "action": "sourceToSeed",
            "status": "queued",
            "createdAt": now_text(),
        }
        write_json(job_path, job)
        thread = threading.Thread(target=run_seed_job, args=(job_id, payload), daemon=True)
        thread.start()
        self.send_json(job, 202)

    def handle_keyword_generate(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        job_id = uuid.uuid4().hex[:12]
        job_path = JOBS_ROOT / f"{job_id}.json"
        job = {
            "ok": True,
            "jobId": job_id,
            "action": "keywordGenerate",
            "status": "queued",
            "createdAt": now_text(),
            "progressPercent": 1,
            "currentStage": "키워드 생성 대기",
        }
        write_json(job_path, job)
        thread = threading.Thread(target=run_keyword_job, args=(job_id, payload), daemon=True)
        thread.start()
        self.send_json(job, 202)

    def handle_job_stop(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        job_id = text_value(payload.get("jobId"))
        if not job_id:
            self.send_json({"ok": False, "error": "jobId required"}, 400)
            return
        stopped = stop_active_job(job_id)
        job_path = JOBS_ROOT / f"{job_id}.json"
        job = read_json(job_path, {"jobId": job_id})
        job.update({
            "ok": True,
            "jobId": job_id,
            "status": "cancelled",
            "finishedAt": now_text(),
            "currentStage": "사용자 중지",
            "stopped": stopped,
        })
        write_json(job_path, job)
        self.send_json({"ok": True, "jobId": job_id, "stopped": stopped})

    def handle_market_upload(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        with ACTIVE_MARKET_UPLOAD_LOCK:
            active_job_id = find_active_market_upload_job()
            if active_job_id:
                self.send_json({
                    "ok": True,
                    "jobId": active_job_id,
                    "deduped": True,
                    "status": "running",
                    "currentStage": "기존 업로드 작업 진행 중",
                }, 202)
                return
            job_id = uuid.uuid4().hex[:12]
            job_path = JOBS_ROOT / f"{job_id}.json"
            job = {
                "ok": True,
                "jobId": job_id,
                "action": "marketUpload",
                "status": "queued",
                "createdAt": now_text(),
                "progressPercent": 1,
                "currentStage": "업로드 대기",
            }
            write_json(job_path, job)
            thread = threading.Thread(target=run_market_upload_job, args=(job_id, payload), daemon=True)
            thread.start()
        self.send_json(job, 202)

    def handle_excel_export(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        export = write_market_export(payload)
        self.send_json({"ok": True, "export": export})

    def handle_workspace_reset(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        cleanup = cleanup_workspace_artifacts(payload)
        self.send_json({"ok": True, "cleanup": cleanup, "seeds": list_seed_summaries()})

    def handle_seed_action(self) -> None:
        payload = json.loads(self.read_body().decode("utf-8", errors="replace") or "{}")
        action = text_value(payload.get("action"))
        seed_path = resolve_seed_path(payload.get("path") or payload.get("name") or "")
        if not seed_path.exists():
            self.send_json({"ok": False, "error": "seed not found"}, 404)
            return
        if action == "delete":
            seed_path.unlink()
            self.send_json({"ok": True, "action": "delete", "seeds": list_seed_summaries()})
            return
        if action == "rename":
            new_name = safe_name(text_value(payload.get("newName")))
            if not new_name.lower().endswith(".webseed.json"):
                new_name = re.sub(r"\.json$", "", new_name, flags=re.IGNORECASE) + ".webseed.json"
            target = (SEED_ROOT / new_name).resolve()
            if not is_within(SEED_ROOT, target):
                raise ValueError("target path outside data/seeds")
            if target.exists() and target != seed_path:
                raise ValueError("same seed name already exists")
            seed_path.rename(target)
            self.send_json({"ok": True, "action": "rename", "seed": seed_summary(target), "seeds": list_seed_summaries()})
            return
        self.send_json({"ok": False, "error": "unknown seed action"}, 400)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=5500)
    parser.add_argument("--host", default="127.0.0.1")
    args = parser.parse_args()

    mimetypes.add_type("text/javascript; charset=utf-8", ".js")
    mimetypes.add_type("text/jsx; charset=utf-8", ".jsx")
    httpd = ThreadingHTTPServer((args.host, args.port), WebOcrHandler)
    print(f"WEBOCRV2 local API server listening on http://{args.host}:{args.port}", flush=True)
    print(f"ROOT={ROOT}", flush=True)
    httpd.serve_forever()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
